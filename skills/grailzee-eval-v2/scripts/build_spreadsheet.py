"""Branded openpyxl spreadsheet output per guide Section 12.2.

Extracted from v1 analyze_report.py:383-598. Adapted for v2's flat
reference dict (no nested 'analysis' key, no 'section' field, no
individual sales). v1's Raw Data sheet is dropped (individual sales
not available in builder inputs).

Sheets: Buy Targets, Trends, Quick Reference.

Usage:
    build_spreadsheet.py <all_results_json> <trends_json> <changes_json>
        <breakouts_json> <watchlist_json> <brands_json>
        <ledger_stats_json> <output_folder>
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SCRIPT_DIR = Path(__file__).resolve().parent
V2_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(V2_ROOT))

from scripts.grailzee_common import OUTPUT_PATH, VARDALUX_COLORS, get_tracer

tracer = get_tracer(__name__)

# Brand colors
RICH_BLACK = VARDALUX_COLORS["rich_black"]
WARM_GOLD = VARDALUX_COLORS["warm_gold"]
DEEP_TEAL = VARDALUX_COLORS["deep_teal"]
WHITE = "FFFFFF"

# Functional colors (not brand; local to spreadsheet)
GREEN_BG = "E8F5E9"
RED_BG = "FFEBEE"
YELLOW_BG = "FFFDE7"
LIGHT_GRAY = "F5F5F5"

HEADERS = [
    "Brand", "Model", "Reference", "Median", "ST%",
    "MAX BUY", "Risk(VG+)", "Signal", "Format", "Floor", "Vol", "Notes",
]

COL_WIDTHS = {
    "A": 12, "B": 24, "C": 22, "D": 14, "E": 8, "F": 14,
    "G": 10, "H": 14, "I": 10, "J": 10, "K": 8, "L": 40,
}


def _styles() -> dict:
    """Return reusable style dict. Extracted from v1 s()."""
    thin = Side(style="thin", color="D9D9D9")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "hfont": Font(name="Arial", bold=True, color=WHITE, size=10),
        "hfill": PatternFill("solid", fgColor=RICH_BLACK),
        "gold": PatternFill("solid", fgColor=WARM_GOLD),
        "teal": PatternFill("solid", fgColor=DEEP_TEAL),
        "green": PatternFill("solid", fgColor=GREEN_BG),
        "red": PatternFill("solid", fgColor=RED_BG),
        "yellow": PatternFill("solid", fgColor=YELLOW_BG),
        "alt": PatternFill("solid", fgColor=LIGHT_GRAY),
        "dfont": Font(name="Arial", size=10),
        "bfont": Font(name="Arial", bold=True, size=10),
        "gfont": Font(name="Arial", bold=True, color=RICH_BLACK, size=11),
        "ffont": Font(name="Arial", size=8, italic=True, color=WARM_GOLD),
    }


def _write_section(ws, row: int, title: str, subtitle: str, bg: str, tc: str, st: dict) -> None:
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = f"  {title}   |   {subtitle}"
    ws[f"A{row}"].font = Font(name="Arial", bold=True, color=tc, size=9)
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=bg)
    ws[f"A{row}"].alignment = st["left"]
    ws.row_dimensions[row].height = 20


def _write_headers(ws, row: int, st: dict) -> None:
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = st["hfont"]
        c.fill = st["hfill"]
        c.alignment = st["center"]
        c.border = st["border"]
    ws.row_dimensions[row].height = 24


def _write_row(ws, row: int, ref_data: dict, offset: int, st: dict) -> None:
    """Write one reference row. ref_data is v2 flat shape (metrics at top level)."""
    is_res = ref_data.get("recommend_reserve", False)
    mb = ref_data.get("max_buy_res") if is_res else ref_data.get("max_buy_nr")
    risk = ref_data.get("risk_nr")
    st_pct = ref_data.get("st_pct")
    vals = [
        ref_data.get("brand", ""),
        ref_data.get("model", ""),
        ref_data.get("reference", ""),
        ref_data.get("median"),
        f"{st_pct:.0%}" if st_pct is not None else "\u2014",
        mb,
        f"{risk:.0f}%" if risk is not None else "\u2014",
        ref_data.get("signal", ""),
        "Reserve" if is_res else "NR",
        ref_data.get("floor"),
        ref_data.get("volume"),
        "",
    ]
    rf = "FFFFFF" if offset % 2 == 0 else LIGHT_GRAY
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = st["dfont"]
        c.fill = PatternFill("solid", fgColor=rf)
        c.border = st["border"]
        c.alignment = st["center"] if ci >= 4 else st["left"]
        if ci == 4:
            c.number_format = "$#,##0"
        elif ci == 6:
            c.font = st["gfont"]
            c.fill = st["gold"]
            c.number_format = "$#,##0"
        elif ci == 10:
            c.number_format = "$#,##0"
        elif ci == 8:
            sig = str(v)
            if "Strong" in sig:
                c.fill = st["green"]
            elif "Reserve" in sig or "Careful" in sig:
                c.fill = st["yellow"]
            elif "Pass" in sig:
                c.fill = st["red"]
        elif ci == 9 and is_res:
            c.fill = st["yellow"]
    ws.row_dimensions[row].height = 22


def build_spreadsheet(
    all_results: dict,
    trends: dict,
    changes: dict,
    breakouts: dict,
    watchlist: dict,
    brands: dict,
    ledger_stats: dict,
    output_folder: str,
) -> str:
    """Produce branded xlsx spreadsheet. Returns output path."""
    wb = openpyxl.Workbook()
    st = _styles()
    ws = wb.active
    ws.title = "Buy Targets"

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # Title (row 1)
    ws.merge_cells("A1:L1")
    ws["A1"] = "VARDALUX COLLECTIONS \u2014 GRAILZEE BUY TARGETS"
    ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=13)
    ws["A1"].fill = st["hfill"]
    ws["A1"].alignment = st["center"]
    ws.row_dimensions[1].height = 28

    # Subtitle (row 2)
    ws.merge_cells("A2:L2")
    ws["A2"] = (
        f'Updated: {datetime.now().strftime("%B %Y")}  |  '
        "Target: 5%/turn, 10%/month  |  NR: $149  |  Reserve: $199"
    )
    ws["A2"].font = Font(name="Arial", italic=True, color=WARM_GOLD, size=9)
    ws["A2"].fill = st["hfill"]
    ws["A2"].alignment = st["center"]

    refs = all_results.get("references", {})
    dj_configs = all_results.get("dj_configs", {})

    # All References section
    row = 4
    _write_section(ws, row, "ALL REFERENCES", "Every reference with 3+ sales scored",
                   GREEN_BG, "1B5E20", st)
    row += 1
    _write_headers(ws, row, st)
    row += 1
    off = 0
    for ref in sorted(refs, key=lambda r: refs[r].get("brand", "")):
        _write_row(ws, row, refs[ref], off, st)
        row += 1
        off += 1

    # DJ Configs section
    if dj_configs:
        row += 1
        _write_section(ws, row, "ROLEX DATEJUST 126300 \u2014 BY CONFIG",
                       "Min 3 sales per config", "E3F2FD", "0D47A1", st)
        row += 1
        _write_headers(ws, row, st)
        row += 1
        off = 0
        for cn in sorted(dj_configs):
            _write_row(ws, row, dj_configs[cn], off, st)
            row += 1
            off += 1

    # Footer
    row += 2
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = (
        "  RULES: Never buy above MAX BUY. Every $100 below = $100 profit. "
        "US inventory only. Risk(VG+) > 40% = Reserve."
    )
    ws[f"A{row}"].font = st["ffont"]
    ws[f"A{row}"].fill = st["hfill"]
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False

    # --- Trends tab ---
    trend_entries = trends.get("trends", [])
    if trend_entries:
        ws2 = wb.create_sheet("Trends")
        t_widths = {
            "A": 12, "B": 22, "C": 18, "D": 12, "E": 12, "F": 10,
            "G": 10, "H": 8, "I": 8, "J": 8, "K": 8, "L": 8, "M": 24,
        }
        for col, w in t_widths.items():
            ws2.column_dimensions[col].width = w

        ws2.merge_cells("A1:M1")
        ws2["A1"] = "TREND COMPARISON"
        ws2["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
        ws2["A1"].fill = st["hfill"]
        ws2["A1"].alignment = st["center"]

        notable = [t for t in trend_entries if t.get("signals")]
        if notable:
            ws2.merge_cells("A2:M2")
            key_text = "  KEY: " + "  |  ".join(
                f"{t['model']}: {t['signal_str']}" for t in notable[:5]
            )
            ws2["A2"] = key_text
            ws2["A2"].font = Font(name="Arial", bold=True, color=WARM_GOLD, size=9)
            ws2["A2"].fill = PatternFill("solid", fgColor=DEEP_TEAL)

        th = [
            "Brand", "Model", "Ref", "Prev Med", "Curr Med", "Chg $", "Chg %",
            "Prev ST", "Curr ST", "Chg ST", "Prev Vol", "Curr Vol", "Signal",
        ]
        for ci, h in enumerate(th, 1):
            c = ws2.cell(row=4, column=ci, value=h)
            c.font = st["hfont"]
            c.fill = st["hfill"]
            c.alignment = st["center"]
            c.border = st["border"]

        for i, t in enumerate(trend_entries):
            r = 5 + i
            vals = [
                t.get("brand", ""), t.get("model", ""), t.get("reference", ""),
                t.get("prev_median"), t.get("curr_median"),
                t.get("med_change"),
                f"{t.get('med_pct', 0):.1f}%",
                f"{t['prev_st']:.0%}" if t.get("prev_st") is not None else "\u2014",
                f"{t['curr_st']:.0%}" if t.get("curr_st") is not None else "\u2014",
                f"{t['st_change']:+.0f}pp" if t.get("st_change") is not None else "\u2014",
                t.get("prev_vol", 0), t.get("curr_vol", 0),
                t.get("signal_str", "Stable"),
            ]
            rf = "FFFFFF" if i % 2 == 0 else LIGHT_GRAY
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(row=r, column=ci, value=v)
                c.font = st["dfont"]
                c.border = st["border"]
                c.fill = PatternFill("solid", fgColor=rf)
                c.alignment = st["center"] if ci >= 4 else st["left"]
                if ci in (4, 5):
                    c.number_format = "$#,##0"
                elif ci == 6:
                    c.number_format = "$#,##0"
                    if isinstance(v, (int, float)):
                        if v > 0:
                            c.fill = st["green"]
                        elif v < 0:
                            c.fill = st["red"]
                elif ci == 13:
                    sv = str(v)
                    if "Cooling" in sv or "Down" in sv:
                        c.fill = st["red"]
                    elif "Momentum" in sv or "Up" in sv:
                        c.fill = st["green"]
                    elif "Reserve" in sv:
                        c.fill = st["yellow"]
        ws2.freeze_panes = "A5"

    # --- Quick Reference tab ---
    ws3 = wb.create_sheet("Quick Reference")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 45
    qr = [
        ("FEES", "", True),
        ("No-Reserve (Branded)", "$149 total", False),
        ("Reserve", "$199 total", False),
        ("", "", False),
        ("RETURN MODEL", "", True),
        ("Monthly target", "10% on deployed capital", False),
        ("Per-trade margin", "5%", False),
        ("Capital cycle", "~14 days", False),
        ("Turns/month", "2", False),
        ("Capital base", "$40,000", False),
        ("", "", False),
        ("FORMAT RULES", "", True),
        ("Default", "NR via Branded Account", False),
        ("Reserve trigger", "Risk(VG+) > 40%", False),
        ("", "", False),
        ("RISK SIGNALS", "", True),
        ("0-10% Strong", "Buy confidently", False),
        ("11-20% Normal", "Fine on clean pieces", False),
        ("21-40% Reserve", "Use Reserve account", False),
        ("41-50% Careful", "Source below Max Buy", False),
        ("50%+ Pass", "Do not source", False),
        ("", "", False),
        ("SOURCING", "", True),
        ("1. Private FB groups", "Best margins", False),
        ("2. Tuesday dealers", "Weekly access", False),
        ("3. eBay BIN", "Occasional", False),
        ("4. Chrono24 US", "Last resort", False),
        ("5. US inventory ONLY", "Tariff avoidance", False),
    ]
    for i, (label, value, is_hdr) in enumerate(qr):
        c1 = ws3.cell(row=i + 1, column=1, value=label)
        c2 = ws3.cell(row=i + 1, column=2, value=value)
        if is_hdr:
            c1.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c1.fill = PatternFill("solid", fgColor=DEEP_TEAL)
            c2.fill = PatternFill("solid", fgColor=DEEP_TEAL)
        else:
            c1.font = st["bfont"]
            c2.font = st["dfont"]

    # Save
    os.makedirs(output_folder, exist_ok=True)
    filename = f"Vardalux_Grailzee_Buy_Targets_{datetime.now().strftime('%B%Y')}.xlsx"
    output_path = os.path.join(output_folder, filename)
    wb.save(output_path)
    return output_path


# --- CLI entry ────────────────────────────────────────────────────────


def run(
    all_results: dict,
    trends: dict,
    changes: dict,
    breakouts: dict,
    watchlist: dict,
    brands: dict,
    ledger_stats: dict,
    output_folder: str,
) -> str:
    """CLI-friendly wrapper."""
    return build_spreadsheet(
        all_results, trends, changes, breakouts,
        watchlist, brands, ledger_stats, output_folder,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("output_folder", help="Directory to write xlsx")
    args = parser.parse_args()

    with tracer.start_as_current_span("build_spreadsheet.run") as span:
        # In practice called from orchestrator with dicts; CLI is for testing
        result_path = run({}, {}, {}, {}, {}, {}, {}, args.output_folder)
        span.set_attribute("output_path", result_path)
        print(result_path)
        return 0


if __name__ == "__main__":
    sys.exit(main())
