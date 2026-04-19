"""Grailzee Pro Excel to normalized CSV converter.

Reads a Grailzee Pro biweekly report workbook, extracts sales data
from 'Auctions Sold' and sell-through rates from 'Top Selling Watches',
joins them by reference number, and writes a normalized CSV.

Usage:
    ingest_report.py <input.xlsx> --output-dir PATH [--overwrite]

Output: JSON on stdout with path, row counts, and warnings.
Exit 0 on success (with or without warnings), non-zero on failure.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
V2_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(V2_ROOT))

from scripts.grailzee_common import get_tracer

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

tracer = get_tracer(__name__)

# Output CSV columns
OUTPUT_COLUMNS = [
    "date_sold", "make", "reference", "title", "condition",
    "papers", "sold_price", "sell_through_pct",
]

# Required columns in Auctions Sold (must all be present in row 1)
REQUIRED_AUCTION_COLUMNS = {"Make", "Model", "Reference Number", "Sold For", "Condition", "Papers"}


# ─── Normalization helpers ────────────────────────────────────────────


def normalize_price(val) -> float | None:
    """Normalize price values from Excel.

    Handles: int, float, str with $ and commas, .0 suffix.
    Returns None for unparseable values.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("$", "").replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normalize_reference(val) -> str | None:
    """Normalize reference number from Excel.

    Handles: int (126300), float (126300.0), str ("A17320").
    Returns None for blank/missing.
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if s.endswith(".0"):
        s = s[:-2]
    return s


def normalize_year(val) -> int | None:
    """Normalize year from Excel.

    Handles: int, float (2025.0), "Unknown", blank.
    Returns None for non-numeric.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip().lower()
    if s in ("", "unknown", "none", "n/a"):
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def normalize_sell_through(val) -> float | None:
    """Normalize sell-through rate from Excel.

    Handles: "23%" -> 0.23, 0.28 -> 0.28, 23 -> 0.23, None -> None.
    """
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip().replace("%", "")
        if not s:
            return None
        try:
            v = float(s)
        except ValueError:
            return None
    elif isinstance(val, (int, float)):
        v = float(val)
    else:
        return None
    # If > 1, treat as percentage (e.g. 23 -> 0.23)
    if v > 1:
        return round(v / 100, 4)
    return round(v, 4)


def normalize_date(val) -> str | None:
    """Normalize date from Excel datetime to YYYY-MM-DD string.

    Handles: datetime objects, date objects, ISO strings.
    Returns None for blank/unparseable.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if hasattr(val, "isoformat"):
        return val.isoformat()[:10]
    s = str(val).strip()
    if not s:
        return None
    # Try ISO parse
    try:
        return datetime.fromisoformat(s).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return None


# ─── Sheet parsers ────────────────────────────────────────────────────


def _build_column_map(ws) -> dict[str, int]:
    """Map header names to column indices (0-based) from row 1."""
    header_map = {}
    for cell in ws[1]:
        if cell.value is not None:
            header_map[str(cell.value).strip()] = cell.column - 1
    return header_map


def parse_auctions_sold(ws) -> tuple[list[dict], list[str]]:
    """Parse the Auctions Sold sheet. Returns (rows, warnings)."""
    warnings: list[str] = []
    col_map = _build_column_map(ws)

    # Check required columns
    actual_cols = set(col_map.keys())
    missing = REQUIRED_AUCTION_COLUMNS - actual_cols
    if missing:
        return [], [f"FATAL: Missing required columns: {sorted(missing)}"]

    extra = actual_cols - {"Sold at", "Auction", "Make", "Model",
                           "Reference Number", "Sold For", "Condition",
                           "Year", "Papers", "Box", "URL"}
    if extra:
        warnings.append(f"Unexpected columns ignored: {sorted(extra)}")

    rows = []
    skipped = 0
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row_cells)

        def get(col_name: str):
            idx = col_map.get(col_name)
            if idx is not None and idx < len(vals):
                return vals[idx]
            return None

        ref = normalize_reference(get("Reference Number"))
        if not ref:
            skipped += 1
            continue

        price = normalize_price(get("Sold For"))
        if price is None:
            skipped += 1
            continue

        date_str = normalize_date(get("Sold at"))
        if not date_str:
            skipped += 1
            continue

        papers_raw = get("Papers")
        papers = str(papers_raw).strip() if papers_raw else ""

        rows.append({
            "date_sold": date_str,
            "make": str(get("Make") or "").strip(),
            "reference": ref,
            "title": str(get("Auction") or "").strip(),
            "condition": str(get("Condition") or "").strip(),
            "papers": papers,
            "sold_price": price,
        })

    if skipped:
        warnings.append(f"{skipped} rows skipped (missing date, price, or reference)")

    return rows, warnings


def parse_top_selling(ws) -> dict[str, float]:
    """Parse Top Selling Watches. Returns {reference: sell_through_pct}."""
    col_map = _build_column_map(ws)
    ref_idx = col_map.get("Reference Number")
    st_idx = col_map.get("Sell-Through Rate")
    if ref_idx is None or st_idx is None:
        return {}

    result = {}
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row_cells)
        if ref_idx >= len(vals) or st_idx >= len(vals):
            continue
        ref = normalize_reference(vals[ref_idx])
        if not ref:
            continue
        st = normalize_sell_through(vals[st_idx])
        if st is not None:
            result[ref] = st
    return result


# ─── Join + write ─────────────────────────────────────────────────────


def _determine_output_name(sales: list[dict], input_path: str) -> str:
    """Determine the output CSV filename from sales dates or file mtime."""
    if sales:
        dates = [s["date_sold"] for s in sales if s.get("date_sold")]
        if dates:
            latest = max(dates)
            return f"grailzee_{latest}.csv"
    # Fallback to file mtime
    try:
        mtime = os.path.getmtime(input_path)
        dt = datetime.fromtimestamp(mtime)
        resolved = dt.strftime("%Y-%m-%d")
        print(
            f"WARNING: no usable sale dates in workbook {input_path!r}; "
            f"falling back to file mtime as report date: {resolved}.",
            file=sys.stderr,
        )
        return f"grailzee_{resolved}.csv"
    except OSError as exc:
        raise ValueError(
            f"Cannot determine report date for {input_path!r}: "
            "no usable sale dates in workbook and file mtime is unavailable. "
            f"Cause: {exc}"
        ) from exc


def ingest(
    input_path: str,
    output_dir: str,
    overwrite: bool = False,
) -> dict:
    """Main ingestion logic. Returns a result dict for JSON output."""
    warnings: list[str] = []

    # Load workbook
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    # Auctions Sold (required)
    as_sheet = None
    for name in ["Auctions Sold", "auctions sold"]:
        if name in wb.sheetnames:
            as_sheet = wb[name]
            break
    if as_sheet is None:
        wb.close()
        raise ValueError("Missing required sheet: 'Auctions Sold'")

    sales, as_warnings = parse_auctions_sold(as_sheet)
    # Check for fatal warnings from column check
    fatal = [w for w in as_warnings if w.startswith("FATAL:")]
    if fatal:
        wb.close()
        raise ValueError(fatal[0])
    warnings.extend(as_warnings)

    if not sales:
        wb.close()
        raise ValueError("Auctions Sold sheet has no usable data rows")

    # Top Selling Watches (optional)
    ts_sheet = None
    for name in ["Top Selling Watches", "top selling watches"]:
        if name in wb.sheetnames:
            ts_sheet = wb[name]
            break

    sell_through_map: dict[str, float] = {}
    top_selling_rows = 0
    if ts_sheet is not None:
        sell_through_map = parse_top_selling(ts_sheet)
        top_selling_rows = len(sell_through_map)
    else:
        warnings.append(
            "Top Selling Watches sheet missing; sell_through_pct will be empty for all rows"
        )

    wb.close()

    # Join sell-through
    joined = 0
    missing_st = 0
    for sale in sales:
        st = sell_through_map.get(sale["reference"])
        if st is not None:
            sale["sell_through_pct"] = st
            joined += 1
        else:
            sale["sell_through_pct"] = ""
            missing_st += 1

    if missing_st and sell_through_map:
        warnings.append(
            f"{missing_st} sales had no matching reference in Top Selling Watches; "
            f"sell_through_pct empty for those rows"
        )

    # Output path
    os.makedirs(output_dir, exist_ok=True)
    csv_name = _determine_output_name(sales, input_path)
    output_path = os.path.join(output_dir, csv_name)

    if os.path.exists(output_path) and not overwrite:
        raise FileExistsError(
            f"Output CSV already exists: {output_path}. Use --overwrite to replace."
        )

    # Write CSV
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for sale in sales:
            writer.writerow({
                "date_sold": sale["date_sold"],
                "make": sale["make"],
                "reference": sale["reference"],
                "title": sale["title"],
                "condition": sale["condition"],
                "papers": sale["papers"],
                "sold_price": sale["sold_price"],
                "sell_through_pct": sale.get("sell_through_pct", ""),
            })

    return {
        "output_csv": output_path,
        "rows_written": len(sales),
        "sheets": {
            "auctions_sold_rows": len(sales),
            "top_selling_rows": top_selling_rows,
            "sell_through_joined": joined,
            "sell_through_missing": missing_st,
        },
        "warnings": warnings,
    }


# ─── CLI ──────────────────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", help="Path to Grailzee Pro Excel workbook")
    parser.add_argument("--output-dir", required=True,
                        help="Output directory for the normalized CSV")
    parser.add_argument("--overwrite", action="store_true",
                        help="Allow overwriting existing CSV")
    args = parser.parse_args()

    with tracer.start_as_current_span("ingest_report.run") as span:
        span.set_attribute("input_path", args.input)

        if not os.path.exists(args.input):
            print(json.dumps({
                "status": "error",
                "error": f"Input file not found: {args.input}",
            }), file=sys.stderr)
            return 1

        try:
            result = ingest(args.input, args.output_dir, args.overwrite)
        except (ValueError, FileExistsError) as exc:
            print(json.dumps({
                "status": "error",
                "error": str(exc),
            }), file=sys.stderr)
            return 1
        except Exception as exc:
            print(json.dumps({
                "status": "error",
                "error": f"Unexpected error: {exc}",
            }), file=sys.stderr)
            return 1

        span.set_attribute("output_path", result["output_csv"])
        span.set_attribute("rows_written", result["rows_written"])
        span.set_attribute("rows_skipped",
                           result["sheets"].get("sell_through_missing", 0))
        span.set_attribute("sell_through_joined",
                           result["sheets"]["sell_through_joined"])
        span.set_attribute("sell_through_missing",
                           result["sheets"]["sell_through_missing"])
        span.set_attribute("warning_count", len(result["warnings"]))

        print(json.dumps(result, indent=2))
        return 0


if __name__ == "__main__":
    sys.exit(main())
