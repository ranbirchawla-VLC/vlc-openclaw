#!/usr/bin/env python3
"""
Vardalux Grailzee Deal Evaluator

Takes brand + reference + purchase price, reads the analysis cache,
and returns a structured recommendation as JSON.

This is the API-level computation layer. OpenClaw calls Claude with
this skill context; Claude runs this script and returns the result.
OpenClaw routes the output to Telegram or wherever it needs to go.

Usage:
    python3 evaluate_deal.py <brand> <reference> <purchase_price> [--cache <path>]

Output: JSON to stdout
"""

import sys, os, json, re, statistics
from datetime import datetime

# ═══ DEFAULTS ═══
GRAILZEE_ROOT = "/Users/ranbirchawla/Library/CloudStorage/GoogleDrive-ranbir.chawla@rnvillc.com/Shared drives/Vardalux Shared Drive/GrailzeeData"
DEFAULT_CACHE_PATH = os.path.join(GRAILZEE_ROOT, "state", "analysis_cache.json")
REPORTS_DIR = os.path.join(GRAILZEE_ROOT, "reports")

NR_FIXED = 149
RES_FIXED = 199
TARGET_MARGIN = 0.05
QUALITY_CONDITIONS = {'very good', 'like new', 'new', 'excellent'}

# Try to import openpyxl for raw report fallback
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Ad budget brackets (from business-model.md)
AD_BUDGETS = [
    (3500, "$37–50"),
    (5000, "$50–100"),
    (10000, "$200–250"),
    (float('inf'), "$250 cap"),
]

# ═══ REFERENCE MATCHING ═══
def normalize_ref(s):
    """Normalize a reference string for matching."""
    s = str(s).strip().upper()
    if s.endswith('.0'): s = s[:-2]
    return s

def strip_ref(s):
    """Strip common prefixes/suffixes for fuzzy matching."""
    s = normalize_ref(s)
    # Strip leading M (Tudor convention)
    if s.startswith('M') and len(s) > 5:
        s = s[1:]
    # Strip trailing -XXXX suffixes
    s = re.sub(r'-\d{4}$', '', s)
    # Remove all separators for comparison
    return s.replace('-', '').replace('.', '').replace(' ', '')

def find_in_cache(cache, brand_input, ref_input):
    """
    Find a reference in the cache. Returns (cache_key, entry) or (None, None).
    
    Matching priority:
    1. Exact reference match
    2. Stripped reference match (M79830RB-0001 matches 79830RB)
    3. Brand + partial reference match
    """
    norm_ref = normalize_ref(ref_input)
    stripped_ref = strip_ref(ref_input)
    brand_upper = brand_input.strip().upper()

    # Pass 1: exact match on reference field
    for key, entry in cache.get("references", {}).items():
        if normalize_ref(entry.get("reference", "")) == norm_ref:
            return key, entry

    # Pass 2: stripped match
    for key, entry in cache.get("references", {}).items():
        cache_stripped = strip_ref(entry.get("reference", ""))
        if cache_stripped == stripped_ref:
            return key, entry
        # Also check alternate_refs
        for alt in entry.get("alternate_refs", []):
            if strip_ref(alt) == stripped_ref:
                return key, entry

    # Pass 3: brand match + substring
    for key, entry in cache.get("references", {}).items():
        if entry.get("brand", "").upper() != brand_upper:
            continue
        cache_stripped = strip_ref(entry.get("reference", ""))
        if stripped_ref in cache_stripped or cache_stripped in stripped_ref:
            return key, entry

    # Pass 4: check DJ configs separately
    for key, entry in cache.get("dj_configs", {}).items():
        cache_stripped = strip_ref(entry.get("reference", ""))
        if stripped_ref in cache_stripped or cache_stripped in stripped_ref:
            return key, entry

    return None, None


def get_ad_budget(median_price):
    """Return recommended ad budget string based on expected sale price."""
    for threshold, budget in AD_BUDGETS:
        if median_price <= threshold:
            return budget
    return AD_BUDGETS[-1][1]


# ═══ RAW REPORT FALLBACK ═══
def find_latest_report(reports_dir=None):
    """Find the most recent Grailzee Pro report file."""
    reports_dir = reports_dir or REPORTS_DIR
    if not os.path.isdir(reports_dir):
        return None
    import glob
    patterns = ["Grailzee_Pro*.xlsx", "Grailzee Pro*.xlsx"]
    found = set()
    for p in patterns:
        for f in glob.glob(os.path.join(reports_dir, p)):
            if not os.path.basename(f).startswith("~"):
                found.add(f)
    if not found:
        return None
    return max(found, key=os.path.getmtime)


def analyze_from_report(brand_input, ref_input, reports_dir=None):
    """
    Parse the raw Grailzee Pro report for a specific reference.
    Returns a cache-shaped entry dict, or None if reference not found.
    
    This is the fallback when a reference isn't in the cache.
    It runs the same math as the full analyzer, just for one reference.
    """
    if not HAS_OPENPYXL:
        return None

    report_path = find_latest_report(reports_dir)
    if not report_path:
        return None

    norm_ref = normalize_ref(ref_input)
    stripped_ref = strip_ref(ref_input)
    brand_upper = brand_input.strip().upper()

    wb = openpyxl.load_workbook(report_path, data_only=True)

    # Find the Auctions Sold sheet
    ws = None
    for name in ['Auctions Sold', 'auctions sold', 'Sold', 'Sales']:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    # Find column mapping
    mapping = {}
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row_vals = [cell.value for cell in ws[row_idx]]
        header_lower = [str(h).lower().strip() if h else "" for h in row_vals]
        fields = {
            'reference': ['reference number', 'reference', 'ref', 'ref.'],
            'price': ['sold for', 'price', 'sale price', 'sold price', 'final price', 'amount'],
            'condition': ['condition', 'cond'],
            'papers': ['papers', 'paper', 'docs'],
            'make': ['make', 'brand'],
            'title': ['auction', 'title', 'auction title', 'listing'],
        }
        for field, keywords in fields.items():
            for i, h in enumerate(header_lower):
                if any(kw == h or kw in h for kw in keywords):
                    mapping[field] = i
                    break
        if 'reference' in mapping and 'price' in mapping:
            header_row = row_idx
            break

    if 'reference' not in mapping or 'price' not in mapping:
        wb.close()
        return None

    # Scan for matching sales
    sales = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        rl = list(row)
        if len(rl) <= max(mapping.values()):
            continue

        ref_val = rl[mapping['reference']]
        price_val = rl[mapping['price']]
        if ref_val is None or price_val is None:
            continue

        ref_str = str(ref_val).strip()
        if ref_str.endswith('.0'):
            ref_str = ref_str[:-2]

        # Match this sale against the requested reference
        sale_norm = normalize_ref(ref_str)
        sale_stripped = strip_ref(ref_str)
        matched = (sale_norm == norm_ref or sale_stripped == stripped_ref
                   or stripped_ref in sale_stripped or sale_stripped in stripped_ref)

        # Also check brand if available
        if not matched and 'make' in mapping:
            sale_brand = str(rl[mapping['make']] or '').strip().upper()
            if sale_brand == brand_upper:
                if stripped_ref in sale_stripped or sale_stripped in stripped_ref:
                    matched = True

        if not matched:
            continue

        try:
            price = float(str(price_val).replace('$', '').replace(',', '').strip())
        except (ValueError, TypeError):
            continue
        if price <= 0:
            continue

        def safe_get(field):
            idx = mapping.get(field)
            if idx is not None and idx < len(rl) and rl[idx] is not None:
                return str(rl[idx]).strip()
            return ''

        sales.append({
            'price': price,
            'condition': safe_get('condition'),
            'papers': safe_get('papers'),
            'title': safe_get('title'),
        })

    wb.close()

    if len(sales) < 2:
        return None

    # ── Run the same math as the full analyzer ──
    prices = [s['price'] for s in sales]
    quality_prices = []
    for s in sales:
        cond = s.get('condition', '').lower().strip()
        papers = s.get('papers', '').lower().strip()
        if (any(q in cond for q in QUALITY_CONDITIONS) and
                papers in ('yes', 'y', 'true', '1', 'included')):
            quality_prices.append(s['price'])

    median = statistics.median(prices)
    max_buy_nr = round((median - NR_FIXED) / (1 + TARGET_MARGIN), -1)
    max_buy_res = round((median - RES_FIXED) / (1 + TARGET_MARGIN), -1)
    breakeven_nr = max_buy_nr + NR_FIXED

    risk_nr = None
    if quality_prices:
        below = sum(1 for p in quality_prices if p < breakeven_nr)
        risk_nr = (below / len(quality_prices)) * 100

    recommend_reserve = risk_nr is not None and risk_nr > 20
    qc = len(quality_prices)

    if risk_nr is None or qc < 3:
        signal = "Low data"
    elif risk_nr <= 10:
        signal = "Strong"
    elif risk_nr <= 20:
        signal = "Normal"
    elif risk_nr <= 30:
        signal = "Reserve"
    elif risk_nr <= 50:
        signal = "Careful"
    else:
        signal = "Pass"

    # Build a cache-shaped entry
    return {
        "brand": brand_input,
        "model": "",
        "reference": ref_input,
        "section": "on_demand",
        "median": median,
        "mean": statistics.mean(prices),
        "floor": min(prices),
        "ceiling": max(prices),
        "volume": len(prices),
        "quality_count": qc,
        "st_pct": None,
        "max_buy_nr": max_buy_nr,
        "max_buy_res": max_buy_res,
        "breakeven_nr": breakeven_nr,
        "breakeven_res": max_buy_res + RES_FIXED,
        "risk_nr": risk_nr,
        "risk_res": None,
        "recommend_reserve": recommend_reserve,
        "signal": signal,
        "profit_nr": median - max_buy_nr - NR_FIXED,
        "profit_res": median - max_buy_res - RES_FIXED,
        "trend_signal": "No history",
        "trend_median_change": 0,
        "trend_median_pct": 0,
        "_source": "raw_report",
        "_report": os.path.basename(report_path),
        "_sale_count": len(sales),
    }


def evaluate(brand, reference, purchase_price, cache_path=None):
    """
    Core evaluation function. Returns a dict with the recommendation.
    
    This is the function Claude calls. It returns structured data,
    not formatted text. The caller (Claude in the OpenClaw context)
    formats the response for the downstream interface.
    """
    if cache_path is None:
        cache_path = DEFAULT_CACHE_PATH

    # ── Load cache ──
    if not os.path.exists(cache_path):
        return {
            "status": "error",
            "error": "no_cache",
            "message": f"No analysis cache found at {cache_path}. Run the full Grailzee analyzer first to generate the cache.",
        }

    with open(cache_path, 'r') as f:
        cache = json.load(f)

    # Check cache freshness
    cache_date = cache.get("generated_at", "unknown")
    cache_report = cache.get("source_report", "unknown")
    schema_version = cache.get("schema_version", 0)

    if schema_version < 1:
        return {
            "status": "error",
            "error": "stale_schema",
            "message": "Analysis cache uses an outdated format. Re-run the full analyzer to regenerate.",
        }

    # ── Find the reference ──
    cache_key, entry = find_in_cache(cache, brand, reference)

    if entry is None:
        # Fallback: parse the raw report for this reference
        entry = analyze_from_report(brand, reference)
        if entry is not None:
            # Found in raw data. Flag that this came from on-demand analysis.
            entry["_on_demand"] = True
        else:
            # Not in cache, not in raw report. Needs web research.
            return {
                "status": "not_found",
                "brand": brand,
                "reference": reference,
                "purchase_price": purchase_price,
                "grailzee": "NEEDS_RESEARCH",
                "rationale": (
                    f"No Grailzee sales data for {brand} {reference}. "
                    f"Not in the core program and not found in the raw report. "
                    f"Research Chrono24 and eBay sold comps to establish a median, "
                    f"then apply the standard formula: MAX BUY = (Median - $149) / 1.05."
                ),
                "research_needed": {
                    "brand": brand,
                    "reference": reference,
                    "purchase_price": purchase_price,
                    "search_queries": [
                        f"{brand} {reference} site:chrono24.com",
                        f"{brand} {reference} sold site:ebay.com",
                        f"{brand} {reference} watchrecon",
                    ],
                    "instructions": (
                        "Find 5+ recent sold prices for this reference in VG+ condition with papers. "
                        "Take the median. Apply: MAX BUY NR = (Median - $149) / 1.05. "
                        "If purchase price is below MAX BUY, it's a buy. "
                        "If fewer than 5 comps exist, flag as insufficient data."
                    ),
                },
                "cache_date": cache_date,
                "cache_report": cache_report,
            }

    # ── Extract metrics ──
    a = entry  # the flattened cache entry
    median = a["median"]
    max_buy_nr = a["max_buy_nr"]
    max_buy_res = a["max_buy_res"]
    risk_nr = a.get("risk_nr")
    signal = a["signal"]
    recommend_reserve = a.get("recommend_reserve", False)
    st_pct = a.get("st_pct")
    volume = a.get("volume", 0)
    floor_price = a.get("floor", 0)
    trend = a.get("trend_signal", "Stable")

    # ── Determine format and effective MAX BUY ──
    if recommend_reserve:
        fmt = "Reserve"
        max_buy = max_buy_res
        fixed_cost = RES_FIXED
    else:
        fmt = "NR"
        max_buy = max_buy_nr
        fixed_cost = NR_FIXED

    # ── Price evaluation ──
    margin_dollars = median - purchase_price - fixed_cost
    margin_pct = (margin_dollars / purchase_price * 100) if purchase_price > 0 else 0
    vs_max_buy = purchase_price - max_buy  # negative = below max buy (good)
    vs_median_pct = ((median - purchase_price) / median * 100) if median > 0 else 0

    # ── Decision logic ──
    if signal in ("Pass",):
        grailzee = "NO"
        reason_prefix = (
            f"Risk is too high on {brand} {a.get('model', reference)}. "
            f"Signal is {signal} with {risk_nr:.0f}% of VG+ sales below breakeven."
        )
    elif purchase_price > max_buy:
        over_by = purchase_price - max_buy
        grailzee = "NO"
        reason_prefix = (
            f"Price is ${over_by:,.0f} over MAX BUY (${max_buy:,.0f}). "
            f"At ${purchase_price:,.0f} the margin is {margin_pct:.1f}%, below the 5% target."
        )
    elif purchase_price > max_buy * 0.98 and signal not in ("Strong",):
        grailzee = "MAYBE"
        reason_prefix = (
            f"Price is near the MAX BUY ceiling (${max_buy:,.0f}) and signal is {signal}. "
            f"Margin at median is {margin_pct:.1f}%. Tight, only worth it on a clean piece with papers."
        )
    elif signal in ("Careful", "Reserve") and not recommend_reserve:
        grailzee = "MAYBE"
        reason_prefix = (
            f"Signal is {signal} ({risk_nr:.0f}% VG+ risk). "
            f"Price works at ${purchase_price:,.0f} but route to Reserve account, not branded NR."
        )
        fmt = "Reserve"
        max_buy = max_buy_res
        fixed_cost = RES_FIXED
        margin_dollars = median - purchase_price - fixed_cost
        margin_pct = (margin_dollars / purchase_price * 100) if purchase_price > 0 else 0
    else:
        grailzee = "YES"
        if purchase_price <= max_buy * 0.90:
            reason_prefix = (
                f"Strong buy. ${purchase_price:,.0f} is well below MAX BUY (${max_buy:,.0f}), "
                f"giving {margin_pct:.1f}% margin at median. Signal is {signal}."
            )
        else:
            reason_prefix = (
                f"Buy works. ${purchase_price:,.0f} is within MAX BUY (${max_buy:,.0f}), "
                f"{margin_pct:.1f}% margin at median. Signal is {signal}."
            )

    # ── Add context to rationale ──
    context_parts = []
    if st_pct is not None:
        context_parts.append(f"Sell-through {st_pct:.0%}")
    if volume:
        context_parts.append(f"{volume} sales in period")
    if trend and trend != "Stable":
        context_parts.append(f"Trending: {trend}")
    context = ". ".join(context_parts) + "." if context_parts else ""

    rationale = f"{reason_prefix} {context}".strip()

    # ── Reserve price suggestion ──
    reserve_price = None
    if fmt == "Reserve":
        # Set reserve at breakeven + small buffer (ensures no loss)
        reserve_price = round(purchase_price + fixed_cost + (purchase_price * 0.02), -1)

    # ── Build response ──
    on_demand = entry.get("_on_demand", False)
    
    # Add data quality note for on-demand analysis
    if on_demand:
        rationale += (f" NOTE: This reference is not in the core program. "
                      f"Analysis based on {volume} raw sales from the report. "
                      f"No sell-through or trend data available.")

    result = {
        "status": "ok",
        "brand": a.get("brand", brand),
        "model": a.get("model", ""),
        "reference": a.get("reference", reference),
        "section": a.get("section", "unknown"),
        "purchase_price": purchase_price,
        "data_source": "raw_report" if on_demand else "cache",

        # The recommendation
        "grailzee": grailzee,
        "format": fmt,
        "reserve_price": reserve_price,
        "ad_budget": get_ad_budget(median),
        "rationale": rationale,

        # Supporting numbers (for anyone who wants to dig in)
        "metrics": {
            "median": median,
            "max_buy": max_buy,
            "floor": floor_price,
            "margin_dollars": round(margin_dollars),
            "margin_pct": round(margin_pct, 1),
            "risk_vg_pct": round(risk_nr, 1) if risk_nr is not None else None,
            "signal": signal,
            "sell_through": f"{st_pct:.0%}" if st_pct else None,
            "volume": volume,
            "trend": trend,
            "vs_max_buy": round(vs_max_buy),
        },

        # Cache metadata
        "cache_date": cache_date,
        "cache_report": cache_report,
    }

    return result


# ═══ CLI ═══
if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("Usage: python3 evaluate_deal.py <brand> <reference> <purchase_price> [--cache <path>]")
        print('Example: python3 evaluate_deal.py Tudor 79830RB 2750')
        sys.exit(1)

    brand = sys.argv[1]
    ref = sys.argv[2]
    try:
        price = float(sys.argv[3].replace('$', '').replace(',', ''))
    except ValueError:
        print(json.dumps({"status": "error", "error": "bad_price", "message": f"Cannot parse price: {sys.argv[3]}"}))
        sys.exit(1)

    cache_path = None
    if '--cache' in sys.argv:
        idx = sys.argv.index('--cache')
        if idx + 1 < len(sys.argv):
            cache_path = sys.argv[idx + 1]

    result = evaluate(brand, ref, price, cache_path)
    print(json.dumps(result, indent=2))

