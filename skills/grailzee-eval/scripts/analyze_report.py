#!/usr/bin/env python3
"""
Grailzee Analyzer — Vardalux Buy Target Engine
Processes Grailzee Pro bi-weekly reports and generates buy targets.

Usage:
    python3 analyze_report.py <folder_path> [--output <output_path>]
"""

import sys, os, glob, json, re, statistics
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)

# Cache writer for OpenClaw deal evaluator integration
try:
    from write_cache import write_cache as save_analysis_cache
    CACHE_AVAILABLE = True
except ImportError:
    CACHE_AVAILABLE = False

# ═══ CONSTANTS ═══
NR_FIXED = 149   # $49 fee + $100 ship
RES_FIXED = 199  # $99 fee + $100 ship
TARGET_MARGIN = 0.05
RISK_RESERVE_THRESHOLD = 20
DISCOVERY_MIN_SALES = 5
QUALITY_CONDITIONS = {'very good', 'like new', 'new', 'excellent'}

# Brand colors
RICH_BLACK = "231F20"
WARM_GOLD = "C9A84C"
DEEP_TEAL = "315159"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
GREEN_BG = "E8F5E9"
RED_BG = "FFEBEE"
YELLOW_BG = "FFFDE7"

# ═══ CORE REFERENCES ═══
CORE_REFERENCES = [
    ("Tudor", "Royal 41mm", ["28600"], "core"),
    ("Tudor", "1926 41mm", ["91650"], "core"),
    ("Tudor", "1926 39mm", ["91550"], "core"),
    ("Tudor", "BB GMT Pepsi", ["79830RB", "M79830RB"], "core"),
    ("Tudor", "BB Heritage Red", ["79230R", "M79230R"], "core"),
    ("Tudor", "BB Heritage Blue", ["79230B", "M79230B"], "core"),
    ("Tudor", "BB 58 Black", ["79030N"], "core"),
    ("Tudor", "BB 58 41mm Black", ["7941A1A0RU"], "core"),
    ("Tudor", "BB 58 GMT Coke", ["7939G1A0NRU"], "core"),
    ("Tudor", "Ranger", ["79950"], "core"),
    ("Omega", "SMD 300M Blue", ["210.30.42.20.03.001"], "core"),
    ("Omega", "SMD 300M Black", ["210.30.42.20.04.001"], "core"),
    ("Breitling", "SO Heritage 42", ["A17320"], "core"),
    ("Breitling", "Navitimer 41 Auto", ["A17326"], "core"),
    ("Breitling", "Navitimer 41 Chrono", ["AB0138241C1A1"], "core"),
    ("Cartier", "Santos 40mm", ["WSSA0030"], "core"),
    ("Rolex", "Datejust 41", ["126300"], "dj"),
    ("Rolex", "Air-King (prev)", ["116900"], "opportunistic"),
    ("Rolex", "Air-King", ["126900"], "opportunistic"),
    ("Rolex", "Submariner Date", ["126610LN"], "opportunistic"),
    ("Omega", "NTTD Titanium", ["210.90.42.20.01.001"], "opportunistic"),
    ("Omega", "NTTD NATO", ["210.92.42.20.01.001"], "opportunistic"),
]

DJ_CONFIGS = {
    "Black/Oyster": (["black"], ["oyster"]),
    "Blue/Jubilee": (["blue"], ["jubilee"]),
    "Blue/Oyster": (["blue"], ["oyster"]),
    "Slate/Jubilee": (["slate"], ["jubilee"]),
    "Slate/Oyster": (["slate"], ["oyster"]),
    "Green": (["green"], None),
    "Wimbledon": (["wimbledon"], None),
    "White/Oyster": (["white"], ["oyster"]),
    "Silver": (["silver"], None),
}

# ═══ FILE DISCOVERY ═══
def get_report_date(filepath):
    """Determine report date from data or filename."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sname in ['Auctions Sold', 'auctions sold']:
            if sname in wb.sheetnames:
                ws = wb[sname]
                for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
                    val = row[0]
                    if val and isinstance(val, datetime):
                        wb.close()
                        return val
                    if val and hasattr(val, 'year'):
                        wb.close()
                        return val
        wb.close()
    except Exception:
        pass
    # Fallback: parse month from filename
    name = os.path.basename(filepath).lower()
    months = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
              'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
              'jan':1,'feb':2,'mar':3,'apr':4,'jun':6,'jul':7,'aug':8,
              'sep':9,'oct':10,'nov':11,'dec':12}
    for mname, mnum in months.items():
        if mname in name:
            years = re.findall(r'20\d\d', name)
            year = int(years[0]) if years else 2026
            return datetime(year, mnum, 1)
    return datetime.fromtimestamp(os.path.getmtime(filepath))

def find_reports(folder):
    """Find all Grailzee Pro report files, sorted oldest to newest."""
    patterns = ["Grailzee_Pro_BiWeekly_Report*.xlsx", "Grailzee_Pro_Bi-Weekly_Report*.xlsx",
                "Grailzee_Pro_Bi_Weekly*.xlsx", "Grailzee_Pro*.xlsx",
                "Grailzee Pro*.xlsx", "Grailzee Pro Bi-Weekly*.xlsx",
                "Grailzee Pro BiWeekly*.xlsx"]
    found = set()
    for pattern in patterns:
        for f in glob.glob(os.path.join(folder, pattern)):
            if not os.path.basename(f).startswith("~"):
                found.add(f)
    return sorted(found, key=get_report_date)

# ═══ DATA PARSING ═══
def find_column_mapping(headers):
    mapping = {}
    header_lower = [str(h).lower().strip() if h else "" for h in headers]
    fields = {
        'date': ['sold at','date','sold date','sale date'],
        'title': ['auction','title','auction title','listing'],
        'make': ['make','brand'],
        'model': ['model'],
        'reference': ['reference number','reference','ref','ref.'],
        'price': ['sold for','price','sale price','sold price','final price','amount'],
        'condition': ['condition','cond'],
        'year': ['year','watch year'],
        'papers': ['papers','paper','docs'],
        'box': ['box'],
        'url': ['url','link'],
    }
    for field, keywords in fields.items():
        for i, h in enumerate(header_lower):
            if any(kw == h or kw in h for kw in keywords):
                mapping[field] = i
                break
    return mapping

def parse_report(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = None
    for name in ['Auctions Sold', 'auctions sold', 'Sold', 'Sales']:
        if name in wb.sheetnames:
            ws = wb[name]; break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    mapping = {}
    header_row = None
    for row_idx in range(1, min(6, ws.max_row+1)):
        row_vals = [cell.value for cell in ws[row_idx]]
        m = find_column_mapping(row_vals)
        if 'reference' in m and 'price' in m:
            header_row = row_idx; mapping = m; break

    if not mapping or 'reference' not in mapping or 'price' not in mapping:
        print(f"  WARNING: Could not parse {os.path.basename(filepath)}")
        wb.close()
        return [], {}

    print(f"  Parsed: {os.path.basename(filepath)} — {list(mapping.keys())}")
    sales = []
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        rl = list(row)
        if len(rl) <= max(mapping.values()): continue
        ref_val = rl[mapping['reference']]
        price_val = rl[mapping['price']]
        if ref_val is None or price_val is None: continue
        ref_str = str(ref_val).strip()
        if ref_str.endswith('.0'): ref_str = ref_str[:-2]
        try:
            price = float(str(price_val).replace('$','').replace(',','').strip())
        except (ValueError, TypeError):
            continue
        if price <= 0: continue
        def safe_get(field):
            idx = mapping.get(field)
            if idx is not None and idx < len(rl) and rl[idx] is not None:
                return str(rl[idx]).strip()
            return ''
        sales.append({
            'reference': ref_str, 'price': price,
            'condition': safe_get('condition'), 'papers': safe_get('papers'),
            'box': safe_get('box'), 'title': safe_get('title'),
            'make': safe_get('make'), 'date': rl[mapping['date']] if mapping.get('date') and mapping['date'] < len(rl) else None,
        })

    # Sell-through from Top Selling Watches
    st_data = {}
    if 'Top Selling Watches' in wb.sheetnames:
        ws2 = wb['Top Selling Watches']
        row1 = [cell.value for cell in ws2[1]]
        m2 = find_column_mapping(row1)
        st_col = None
        for i, h in enumerate(str(v).lower() for v in row1):
            if 'sell' in h and 'through' in h or 'sell-through' in h:
                st_col = i; break
        if 'reference' in m2 and st_col is not None:
            for row in ws2.iter_rows(min_row=2, values_only=True):
                rl = list(row)
                if m2['reference'] < len(rl) and st_col < len(rl):
                    ref = rl[m2['reference']]
                    st = rl[st_col]
                    if ref and st:
                        ref_str = str(ref).strip()
                        if ref_str.endswith('.0'): ref_str = ref_str[:-2]
                        try:
                            st_val = float(str(st).replace('%','').strip())
                            if st_val > 1: st_val = st_val / 100
                            st_data[ref_str] = st_val
                        except (ValueError, TypeError):
                            pass
    wb.close()
    return sales, st_data

# ═══ REFERENCE MATCHING ═══
def normalize_ref(s):
    s = str(s).strip().upper()
    if s.endswith('.0'): s = s[:-2]
    return s

def match_reference(sale_ref, core_patterns):
    norm = normalize_ref(sale_ref)
    for p in core_patterns:
        np = normalize_ref(p)
        if np in norm or norm in np: return True
        cn = norm.replace('-','').replace('.','').replace(' ','')
        cp = np.replace('-','').replace('.','').replace(' ','')
        if cp in cn or cn in cp: return True
    return False

def classify_dj_config(title):
    tl = title.lower()
    for cfg, (dial_kw, bracelet_kw) in DJ_CONFIGS.items():
        if any(k in tl for k in dial_kw):
            if bracelet_kw is None: return cfg
            if any(k in tl for k in bracelet_kw): return cfg
    return "Other"

# ═══ ANALYSIS ENGINE ═══
def is_quality_sale(sale):
    cond = sale.get('condition','').lower().strip()
    papers = sale.get('papers','').lower().strip()
    return any(q in cond for q in QUALITY_CONDITIONS) and papers in ('yes','y','true','1','included')

def calc_risk(quality_prices, breakeven):
    if not quality_prices: return None
    below = sum(1 for p in quality_prices if p < breakeven)
    return (below / len(quality_prices)) * 100

def analyze_reference(sales, st_pct=None):
    if not sales: return None
    prices = [s['price'] for s in sales]
    quality_prices = [s['price'] for s in sales if is_quality_sale(s)]
    median = statistics.median(prices)

    max_buy_nr = round((median - NR_FIXED) / (1 + TARGET_MARGIN), -1)
    breakeven_nr = max_buy_nr + NR_FIXED
    risk_nr = calc_risk(quality_prices, breakeven_nr)

    max_buy_res = round((median - RES_FIXED) / (1 + TARGET_MARGIN), -1)
    breakeven_res = max_buy_res + RES_FIXED
    risk_res = calc_risk(quality_prices, breakeven_res)

    recommend_reserve = risk_nr is not None and risk_nr > RISK_RESERVE_THRESHOLD
    qc = len(quality_prices)

    if risk_nr is None or qc < 3: signal = "Low data"
    elif risk_nr <= 10: signal = "Strong"
    elif risk_nr <= 20: signal = "Normal"
    elif risk_nr <= 30: signal = "Reserve"
    elif risk_nr <= 50: signal = "Careful"
    else: signal = "Pass"

    return {
        'median': median, 'mean': statistics.mean(prices),
        'floor': min(prices), 'ceiling': max(prices), 'volume': len(prices),
        'st_pct': st_pct, 'quality_count': qc,
        'max_buy_nr': max_buy_nr, 'max_buy_res': max_buy_res,
        'breakeven_nr': breakeven_nr, 'breakeven_res': breakeven_res,
        'risk_nr': risk_nr, 'risk_res': risk_res,
        'profit_nr': median - max_buy_nr - NR_FIXED,
        'profit_res': median - max_buy_res - RES_FIXED,
        'recommend_reserve': recommend_reserve, 'signal': signal,
    }

def match_core_sales(sales, st_data):
    by_ref = defaultdict(list)
    for s in sales:
        by_ref[normalize_ref(s['reference'])].append(s)

    results = {}
    matched = set()
    for brand, model, patterns, section in CORE_REFERENCES:
        key = f"{brand}|{model}"
        msales = []
        st_val = None
        for ref_group, gsales in by_ref.items():
            if match_reference(ref_group, patterns):
                msales.extend(gsales)
                matched.add(ref_group)
                for p in patterns:
                    for sr, sv in st_data.items():
                        if normalize_ref(sr) == normalize_ref(p):
                            st_val = sv
        if msales:
            a = analyze_reference(msales, st_val)
            if a:
                results[key] = {'brand':brand,'model':model,'reference':patterns[0],'section':section,'analysis':a,'sales':msales}

    # DJ configs
    djk = "Rolex|Datejust 41"
    if djk in results:
        configs = defaultdict(list)
        for s in results[djk]['sales']:
            configs[classify_dj_config(s.get('title',''))].append(s)
        dj_cfgs = {}
        for cn, cs in configs.items():
            if len(cs) >= 3:
                a = analyze_reference(cs)
                if a: dj_cfgs[cn] = {'brand':'Rolex','model':f'DJ 41 {cn}','reference':'126300','section':'dj_config','analysis':a,'sales':cs}
        results['_dj_configs'] = dj_cfgs

    # Dynamic discovery
    discoveries = []
    for ref, rsales in by_ref.items():
        if ref in matched or len(rsales) < DISCOVERY_MIN_SALES: continue
        a = analyze_reference(rsales)
        if a:
            discoveries.append({'brand':rsales[0].get('make','?'),'model':rsales[0].get('title','')[:40],
                                'reference':ref,'analysis':a,'count':len(rsales)})
    discoveries.sort(key=lambda d: d['count'], reverse=True)

    return results, matched, discoveries

# ═══ TREND COMPARISON ═══
def compare_periods(curr, prev):
    trends = []
    for key, c in curr.items():
        if key.startswith('_') or key not in prev: continue
        ca, pa = c['analysis'], prev[key]['analysis']
        mc = ca['median'] - pa['median']
        mp = (mc / pa['median'] * 100) if pa['median'] else 0
        stc = None
        if ca['st_pct'] is not None and pa['st_pct'] is not None:
            stc = (ca['st_pct'] - pa['st_pct']) * 100
        signals = []
        if mp <= -5: signals.append("Cooling")
        elif mp >= 10: signals.append("Momentum")
        if stc is not None:
            if stc >= 10: signals.append("Demand Up")
            elif stc <= -10: signals.append("Demand Down")
        pr, cr = pa.get('risk_nr'), ca.get('risk_nr')
        if pr is not None and cr is not None:
            if pr <= 20 and cr > 20: signals.append("Now Reserve")
            elif pr > 20 and cr <= 20: signals.append("Now NR")
        trends.append({
            'brand':c['brand'],'model':c['model'],'reference':c['reference'],
            'prev_median':pa['median'],'curr_median':ca['median'],
            'med_change':mc,'med_pct':mp,
            'prev_st':pa['st_pct'],'curr_st':ca['st_pct'],'st_change':stc,
            'prev_vol':pa['volume'],'curr_vol':ca['volume'],
            'floor_change':ca['floor']-pa['floor'],
            'signals':signals,'signal_str':' | '.join(signals) if signals else 'Stable',
        })
    return trends

# ═══ SPREADSHEET OUTPUT ═══
def s():
    thin = Side(style='thin', color='D9D9D9')
    return {
        'border': Border(left=thin,right=thin,top=thin,bottom=thin),
        'center': Alignment(horizontal='center',vertical='center',wrap_text=True),
        'left': Alignment(horizontal='left',vertical='center',wrap_text=True),
        'hfont': Font(name='Arial',bold=True,color=WHITE,size=10),
        'hfill': PatternFill('solid',fgColor=RICH_BLACK),
        'gold': PatternFill('solid',fgColor=WARM_GOLD),
        'teal': PatternFill('solid',fgColor=DEEP_TEAL),
        'green': PatternFill('solid',fgColor=GREEN_BG),
        'red': PatternFill('solid',fgColor=RED_BG),
        'yellow': PatternFill('solid',fgColor=YELLOW_BG),
        'alt': PatternFill('solid',fgColor=LIGHT_GRAY),
        'dfont': Font(name='Arial',size=10),
        'bfont': Font(name='Arial',bold=True,size=10),
        'gfont': Font(name='Arial',bold=True,color=RICH_BLACK,size=11),
        'nfont': Font(name='Arial',size=9,italic=True,color='666666'),
        'ffont': Font(name='Arial',size=8,italic=True,color=WARM_GOLD),
    }

HEADERS = ['Brand','Model','Reference','Median','ST%','MAX BUY','Risk(VG+)','Signal','Format','Floor','Vol','Notes']

def write_section(ws, row, title, subtitle, bg, tc, st):
    ws.merge_cells(f'A{row}:L{row}')
    ws[f'A{row}'] = f'  {title}   |   {subtitle}'
    ws[f'A{row}'].font = Font(name='Arial',bold=True,color=tc,size=9)
    ws[f'A{row}'].fill = PatternFill('solid',fgColor=bg)
    ws[f'A{row}'].alignment = st['left']
    ws.row_dimensions[row].height = 20

def write_hdrs(ws, row, st):
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = st['hfont']; c.fill = st['hfill']; c.alignment = st['center']; c.border = st['border']
    ws.row_dimensions[row].height = 24

def write_row(ws, row, data, offset, st):
    a = data['analysis']
    is_res = a['recommend_reserve']
    mb = a['max_buy_res'] if is_res else a['max_buy_nr']
    risk = a['risk_nr']
    vals = [data['brand'], data['model'], data['reference'], a['median'],
            f"{a['st_pct']:.0%}" if a['st_pct'] else "—", mb,
            f"{risk:.0f}%" if risk is not None else "—", a['signal'],
            "Reserve" if is_res else "NR", a['floor'], a['volume'], '']
    rf = 'FFFFFF' if offset % 2 == 0 else LIGHT_GRAY
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = st['dfont']; c.fill = PatternFill('solid',fgColor=rf)
        c.border = st['border']; c.alignment = st['center'] if ci >= 4 else st['left']
        if ci == 4: c.number_format = '$#,##0'
        elif ci == 6:
            c.font = st['gfont']; c.fill = st['gold']; c.number_format = '$#,##0'
        elif ci == 10: c.number_format = '$#,##0'
        elif ci == 8:
            if 'Strong' in str(v): c.fill = st['green']
            elif 'Reserve' in str(v) or 'Careful' in str(v): c.fill = st['yellow']
            elif 'Pass' in str(v): c.fill = st['red']
        elif ci == 9 and is_res: c.fill = st['yellow']
    ws.row_dimensions[row].height = 22

def build_spreadsheet(results, trends, discoveries, output_path):
    wb = openpyxl.Workbook()
    st = s()
    ws = wb.active; ws.title = "Buy Targets"

    for col, w in {'A':12,'B':24,'C':22,'D':14,'E':8,'F':14,'G':10,'H':14,'I':10,'J':10,'K':8,'L':40}.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = 'VARDALUX COLLECTIONS — GRAILZEE BUY TARGETS'
    ws['A1'].font = Font(name='Arial',bold=True,color=WHITE,size=13)
    ws['A1'].fill = st['hfill']; ws['A1'].alignment = st['center']
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:L2')
    ws['A2'] = f'Updated: {datetime.now().strftime("%B %Y")}  |  Target: 5%/turn, 10%/month  |  NR: $149  |  Reserve: $199'
    ws['A2'].font = Font(name='Arial',italic=True,color=WARM_GOLD,size=9)
    ws['A2'].fill = st['hfill']; ws['A2'].alignment = st['center']

    row = 4
    # Core
    write_section(ws, row, 'CORE NR REFERENCES', 'Branded Account — $149 fixed cost', GREEN_BG, '1B5E20', st)
    row += 1; write_hdrs(ws, row, st); row += 1
    off = 0
    for k, d in results.items():
        if k.startswith('_') or d.get('section') not in ('core',): continue
        write_row(ws, row, d, off, st); row += 1; off += 1

    # DJ configs
    if '_dj_configs' in results and results['_dj_configs']:
        row += 1
        write_section(ws, row, 'ROLEX DATEJUST 126300 — BY CONFIG', 'Min 3 sales per config', 'E3F2FD', '0D47A1', st)
        row += 1; write_hdrs(ws, row, st); row += 1
        off = 0
        for cn, d in sorted(results['_dj_configs'].items()):
            write_row(ws, row, d, off, st); row += 1; off += 1

    # Opportunistic
    row += 1
    write_section(ws, row, 'OPPORTUNISTIC / HOME RUN', 'Source on sight — higher price points', 'FFF3E0', 'E65100', st)
    row += 1; write_hdrs(ws, row, st); row += 1
    off = 0
    for k, d in results.items():
        if k.startswith('_') or d.get('section') != 'opportunistic': continue
        write_row(ws, row, d, off, st); row += 1; off += 1

    # Footer
    row += 2
    ws.merge_cells(f'A{row}:L{row}')
    ws[f'A{row}'] = '  RULES: Never buy above MAX BUY. Every $100 below = $100 profit. US inventory only. Risk(VG+) > 20% = Reserve.'
    ws[f'A{row}'].font = st['ffont']; ws[f'A{row}'].fill = st['hfill']
    ws.freeze_panes = 'A6'; ws.sheet_view.showGridLines = False

    # ═══ TRENDS TAB ═══
    if trends:
        ws2 = wb.create_sheet("Trends")
        for col, w in {'A':12,'B':22,'C':18,'D':12,'E':12,'F':10,'G':10,'H':8,'I':8,'J':8,'K':8,'L':8,'M':24}.items():
            ws2.column_dimensions[col].width = w
        ws2.merge_cells('A1:M1')
        ws2['A1'] = 'TREND COMPARISON'; ws2['A1'].font = Font(name='Arial',bold=True,color=WHITE,size=12)
        ws2['A1'].fill = st['hfill']; ws2['A1'].alignment = st['center']

        notable = [t for t in trends if t['signals']]
        if notable:
            ws2.merge_cells('A2:M2')
            ws2['A2'] = '  KEY: ' + '  |  '.join(f"{t['model']}: {t['signal_str']}" for t in notable[:5])
            ws2['A2'].font = Font(name='Arial',bold=True,color=WARM_GOLD,size=9)
            ws2['A2'].fill = PatternFill('solid',fgColor=DEEP_TEAL)

        th = ['Brand','Model','Ref','Prev Med','Curr Med','Chg $','Chg %','Prev ST','Curr ST','Chg ST','Prev Vol','Curr Vol','Signal']
        for ci, h in enumerate(th, 1):
            c = ws2.cell(row=4, column=ci, value=h); c.font = st['hfont']; c.fill = st['hfill']; c.alignment = st['center']; c.border = st['border']
        for i, t in enumerate(trends):
            r = 5 + i
            vals = [t['brand'],t['model'],t['reference'],t['prev_median'],t['curr_median'],
                    t['med_change'],f"{t['med_pct']:.1f}%",
                    f"{t['prev_st']:.0%}" if t['prev_st'] else '—',
                    f"{t['curr_st']:.0%}" if t['curr_st'] else '—',
                    f"{t['st_change']:+.0f}pp" if t['st_change'] else '—',
                    t['prev_vol'],t['curr_vol'],t['signal_str']]
            rf = 'FFFFFF' if i % 2 == 0 else LIGHT_GRAY
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(row=r, column=ci, value=v); c.font = st['dfont']; c.border = st['border']
                c.fill = PatternFill('solid',fgColor=rf); c.alignment = st['center'] if ci >= 4 else st['left']
                if ci in (4,5): c.number_format = '$#,##0'
                elif ci == 6:
                    c.number_format = '$#,##0'
                    if isinstance(v,(int,float)):
                        c.fill = st['green'] if v > 0 else st['red'] if v < 0 else PatternFill('solid',fgColor=rf)
                elif ci == 13:
                    if 'Cooling' in str(v) or 'Down' in str(v): c.fill = st['red']
                    elif 'Momentum' in str(v) or 'Up' in str(v): c.fill = st['green']
                    elif 'Reserve' in str(v): c.fill = st['yellow']
        ws2.freeze_panes = 'A5'

    # ═══ QUICK REFERENCE TAB ═══
    ws3 = wb.create_sheet("Quick Reference")
    ws3.column_dimensions['A'].width = 30; ws3.column_dimensions['B'].width = 45
    qr = [
        ('FEES','',True), ('No-Reserve (Branded)',f'$149 total',False), ('Reserve',f'$199 total',False),
        ('','',False), ('RETURN MODEL','',True), ('Monthly target','10% on deployed capital',False),
        ('Per-trade margin','5%',False), ('Capital cycle','~14 days',False), ('Turns/month','2',False),
        ('Capital base','$40,000',False),
        ('','',False), ('FORMAT RULES','',True), ('Default','NR via Branded Account',False),
        ('Reserve trigger','Risk(VG+) > 20%',False),
        ('','',False), ('RISK SIGNALS','',True),
        ('0-10% Strong','Buy confidently',False), ('11-20% Normal','Fine on clean pieces',False),
        ('21-30% Reserve','Use Reserve account',False), ('31-50% Careful','Source below Max Buy',False),
        ('50%+ Pass','Do not source',False),
        ('','',False), ('SOURCING','',True),
        ('1. Private FB groups','Best margins',False), ('2. Tuesday dealers','Weekly access',False),
        ('3. eBay BIN','Occasional',False), ('4. Chrono24 US','Last resort',False),
        ('5. US inventory ONLY','Tariff avoidance',False),
    ]
    for i, (label, value, is_hdr) in enumerate(qr):
        c1 = ws3.cell(row=i+1, column=1, value=label); c2 = ws3.cell(row=i+1, column=2, value=value)
        if is_hdr:
            c1.font = Font(name='Arial',bold=True,color=WHITE,size=10); c1.fill = PatternFill('solid',fgColor=DEEP_TEAL)
            c2.fill = PatternFill('solid',fgColor=DEEP_TEAL)
        else:
            c1.font = st['bfont']; c2.font = st['dfont']

    # ═══ RAW DATA TAB ═══
    ws4 = wb.create_sheet("Raw Data")
    rh = ['Brand','Model','Ref','Price','Condition','Papers','Box','Title','Date']
    for ci, h in enumerate(rh, 1):
        c = ws4.cell(row=1, column=ci, value=h); c.font = st['hfont']; c.fill = st['hfill']
    ws4.column_dimensions['D'].width = 12; ws4.column_dimensions['H'].width = 50
    r = 2
    for k, d in sorted(results.items()):
        if k.startswith('_'):
            if k == '_dj_configs':
                for cn, cd in d.items():
                    for sl in sorted(cd['sales'], key=lambda x: -x['price']):
                        for ci, v in enumerate([cd['brand'],cd['model'],cd['reference'],sl['price'],
                                                sl.get('condition',''),sl.get('papers',''),sl.get('box',''),
                                                sl.get('title',''),str(sl.get('date',''))], 1):
                            c = ws4.cell(row=r,column=ci,value=v); c.font = st['dfont']
                            if ci == 4: c.number_format = '$#,##0'
                        r += 1
            continue
        for sl in sorted(d['sales'], key=lambda x: -x['price']):
            for ci, v in enumerate([d['brand'],d['model'],d['reference'],sl['price'],
                                    sl.get('condition',''),sl.get('papers',''),sl.get('box',''),
                                    sl.get('title',''),str(sl.get('date',''))], 1):
                c = ws4.cell(row=r,column=ci,value=v); c.font = st['dfont']
                if ci == 4: c.number_format = '$#,##0'
            r += 1
    ws4.freeze_panes = 'A2'; ws4.auto_filter.ref = f'A1:I{r-1}'

    wb.save(output_path)
    print(f"\n  Saved: {output_path}")
    return output_path

# ═══ MARKDOWN SUMMARY ═══
def build_summary(results, trends, discoveries, source_file, total_sales, output_dir):
    """Generate a readable markdown analysis summary alongside the spreadsheet."""
    lines = []
    now = datetime.now().strftime("%B %d, %Y")

    lines.append(f"# Vardalux Grailzee Analysis — {now}")
    lines.append(f"Source: {source_file} | {total_sales:,} total sales analyzed\n")

    # ── Headlines ──
    core = {k:v for k,v in results.items() if not k.startswith('_')}
    strong = [v for v in core.values() if v['analysis']['signal'] == 'Strong']
    reserve = [v for v in core.values() if v['analysis']['recommend_reserve']]
    nr_safe = [v for v in core.values() if not v['analysis']['recommend_reserve'] and v['analysis']['signal'] in ('Strong','Normal')]

    lines.append("## Market Snapshot\n")
    lines.append(f"{len(core)} references tracked across the core program. "
                 f"{len(strong)} showing Strong signal, {len(nr_safe)} safe for NR, "
                 f"{len(reserve)} routing to Reserve this cycle.")

    if trends:
        rising = [t for t in trends if t['med_pct'] >= 5]
        falling = [t for t in trends if t['med_pct'] <= -5]
        if rising:
            names = ", ".join(t['model'] for t in rising[:4])
            lines.append(f"Rising: {names}.")
        if falling:
            names = ", ".join(t['model'] for t in falling[:4])
            lines.append(f"Softening: {names}.")
        if not rising and not falling:
            lines.append("Market is broadly stable period over period.")
    else:
        lines.append("No prior report for trend comparison. Trends will populate on the next run.")

    # ── Buy Targets: NR Safe ──
    lines.append("\n## NR Buy Targets (Branded Account)\n")
    lines.append("| Reference | Model | MAX BUY | Signal | ST% | Vol | Trend |")
    lines.append("|-----------|-------|---------|--------|-----|-----|-------|")
    for v in sorted(nr_safe, key=lambda x: x['brand']):
        a = v['analysis']
        st = f"{a['st_pct']:.0%}" if a['st_pct'] else "—"
        trend_match = next((t for t in trends if t['model'] == v['model']), None)
        trend_str = trend_match['signal_str'] if trend_match else "—"
        lines.append(f"| {v['reference']} | {v['brand']} {v['model']} | "
                     f"**${a['max_buy_nr']:,.0f}** | {a['signal']} | {st} | "
                     f"{a['volume']} | {trend_str} |")

    # ── Reserve Candidates ──
    if reserve:
        lines.append("\n## Reserve Candidates\n")
        lines.append("| Reference | Model | MAX BUY (Res) | Risk VG+ | ST% | Vol |")
        lines.append("|-----------|-------|---------------|----------|-----|-----|")
        for v in sorted(reserve, key=lambda x: x['brand']):
            a = v['analysis']
            st = f"{a['st_pct']:.0%}" if a['st_pct'] else "—"
            risk = f"{a['risk_nr']:.0f}%" if a['risk_nr'] is not None else "—"
            lines.append(f"| {v['reference']} | {v['brand']} {v['model']} | "
                         f"**${a['max_buy_res']:,.0f}** | {risk} | {st} | {a['volume']} |")

    # ── DJ Configs ──
    dj_cfgs = results.get('_dj_configs', {})
    if dj_cfgs:
        lines.append("\n## Datejust 126300 by Configuration\n")
        lines.append("| Config | MAX BUY | Signal | Risk VG+ | Vol |")
        lines.append("|--------|---------|--------|----------|-----|")
        for cn, d in sorted(dj_cfgs.items(), key=lambda x: -(x[1]['analysis'].get('max_buy_nr',0))):
            a = d['analysis']
            mb = a['max_buy_res'] if a['recommend_reserve'] else a['max_buy_nr']
            risk = f"{a['risk_nr']:.0f}%" if a['risk_nr'] is not None else "—"
            fmt_note = " (Res)" if a['recommend_reserve'] else ""
            lines.append(f"| {cn} | **${mb:,.0f}**{fmt_note} | {a['signal']} | {risk} | {a['volume']} |")

    # ── Trend Movers ──
    if trends:
        notable = [t for t in trends if t['signals']]
        if notable:
            lines.append("\n## Notable Moves\n")
            for t in notable:
                direction = "up" if t['med_change'] > 0 else "down"
                lines.append(f"**{t['brand']} {t['model']}**: Median {direction} "
                             f"${abs(t['med_change']):,.0f} ({t['med_pct']:+.1f}%). "
                             f"{t['signal_str']}.")

    # ── Top Discoveries ──
    if discoveries:
        top_disc = [d for d in discoveries[:10] if d['analysis']['signal'] in ('Strong','Normal')]
        if top_disc:
            lines.append("\n## Discovered References (Not in Core Program)\n")
            lines.append("| Reference | Brand | Median | Signal | Sales | Note |")
            lines.append("|-----------|-------|--------|--------|-------|------|")
            for d in top_disc[:8]:
                a = d['analysis']
                lines.append(f"| {d['reference']} | {d['brand']} | "
                             f"${a['median']:,.0f} | {a['signal']} | {d['count']} | "
                             f"Review for core program |")
        lines.append(f"\n{len(discoveries)} total references discovered with 5+ sales. "
                     f"Only Strong/Normal signals shown above.")

    # ── Projection ──
    lines.append("\n## Capital Projection\n")
    lines.append("$40,000 deployed at 5% margin, 2 turns per month = $4,000/month target.")
    thin_refs = [v for v in core.values() if v['analysis']['volume'] < 5]
    if thin_refs:
        names = ", ".join(f"{v['brand']} {v['model']}" for v in thin_refs)
        lines.append(f"Thin data ({names}) may not support reliable 14-day cycles.")

    lines.append(f"\n---\n*Generated {now} from {source_file}*")

    # Write the file
    summary_path = os.path.join(output_dir,
                                f"Vardalux_Grailzee_Analysis_{datetime.now().strftime('%B%Y')}.md")
    with open(summary_path, 'w') as f:
        f.write('\n'.join(lines))
    print(f"  Summary: {summary_path}")
    return summary_path


# ═══ SOURCING BRIEF (AGENT OUTPUT) ═══
def build_sourcing_brief(results, trends, discoveries, source_file, output_dir):
    """
    Generate a structured sourcing brief for automated agents.
    
    This file tells a sourcing agent: what to look for, where, at what price,
    and how urgently. It powers web scraping (eBay, Chrono24, Facebook),
    WhatsApp group monitoring, and dealer chat scanning.
    
    Output: JSON file in GrailzeeData/state/ (machine-readable)
           + markdown brief in GrailzeeData/output/ (human-reviewable)
    """
    now = datetime.now()
    
    # ── Build the target list ──
    targets = []
    
    for key, data in results.items():
        if key.startswith('_'):
            continue
        a = data['analysis']
        if a['signal'] in ('Pass', 'Low data'):
            continue
        
        is_reserve = a['recommend_reserve']
        max_buy = a['max_buy_res'] if is_reserve else a['max_buy_nr']
        sweet_spot = round(max_buy * 0.90, -1)
        
        # Trend context
        trend_match = next((t for t in trends if
                           t.get('brand') == data.get('brand') and
                           t.get('model') == data.get('model')), None)
        trend_str = trend_match['signal_str'] if trend_match else "Stable"
        trend_pct = trend_match['med_pct'] if trend_match else 0
        
        # Priority scoring
        priority_score = 0
        if a['signal'] == 'Strong': priority_score += 3
        elif a['signal'] == 'Normal': priority_score += 2
        elif a['signal'] in ('Reserve', 'Careful'): priority_score += 1
        if trend_pct >= 5: priority_score += 2
        elif trend_pct >= 0: priority_score += 1
        if a['volume'] >= 15: priority_score += 1
        st = a.get('st_pct')
        if st and st >= 0.60: priority_score += 1
        
        if priority_score >= 6: priority = "HIGH"
        elif priority_score >= 4: priority = "MEDIUM"
        else: priority = "LOW"
        
        # Search terms
        brand = data['brand']
        model = data['model']
        ref = data['reference']
        search_terms = [f"{brand} {ref}", f"{brand} {model}"]
        
        if brand == 'Tudor':
            if not ref.startswith('M'):
                search_terms.append(f"Tudor M{ref}")
            else:
                search_terms.append(f"Tudor {ref[1:]}")
        elif brand == 'Omega' and '.' in ref:
            search_terms.append(f"Omega {ref.replace('.','')}")
        
        notes = []
        if is_reserve:
            notes.append(f"Route to Reserve account. Risk at {a['risk_nr']:.0f}%.")
        if trend_pct <= -5:
            notes.append("Softening. Buy at sweet spot or below, not at MAX.")
        if a['volume'] < 8:
            notes.append("Low volume. Fewer opportunities but less competition.")
        if st and st >= 0.75:
            notes.append(f"High sell-through ({st:.0%}). Moves fast on Grailzee.")
        
        targets.append({
            "brand": brand, "model": model, "reference": ref,
            "section": data.get('section', 'core'),
            "priority": priority, "priority_score": priority_score,
            "max_buy": max_buy, "sweet_spot": sweet_spot,
            "median": a['median'], "floor": a['floor'],
            "format": "Reserve" if is_reserve else "NR",
            "signal": a['signal'],
            "risk_vg_pct": round(a['risk_nr'], 1) if a['risk_nr'] is not None else None,
            "volume": a['volume'],
            "sell_through": f"{st:.0%}" if st else None,
            "trend": trend_str, "trend_pct": round(trend_pct, 1),
            "search_terms": search_terms,
            "condition_filter": ["Excellent", "Like New", "Very Good", "BNIB"],
            "papers_required": True,
            "action": "auto_evaluate" if priority in ("HIGH", "MEDIUM") else "flag_for_review",
            "notes": " ".join(notes),
        })
    
    # ── DJ 126300 configs ──
    dj_cfgs = results.get('_dj_configs', {})
    if isinstance(dj_cfgs, dict):
        for cfg_name, data in dj_cfgs.items():
            a = data['analysis']
            if a['signal'] in ('Pass', 'Low data'):
                continue
            is_reserve = a['recommend_reserve']
            max_buy = a['max_buy_res'] if is_reserve else a['max_buy_nr']
            sweet_spot = round(max_buy * 0.90, -1)
            parts = cfg_name.split('/')
            dial = parts[0].lower() if parts else ""
            bracelet = parts[1].lower() if len(parts) > 1 else ""
            search_terms = [f"Rolex Datejust 126300 {dial}"]
            if bracelet:
                search_terms.append(f"Rolex 126300 {dial} {bracelet}")
                search_terms.append(f"Rolex DJ 41 {dial} {bracelet}")
            ps = 2 if a['signal'] in ('Strong', 'Normal') else 1
            if a['volume'] >= 10: ps += 1
            targets.append({
                "brand": "Rolex", "model": f"Datejust 41 {cfg_name}",
                "reference": "126300", "config": cfg_name,
                "section": "dj_config",
                "priority": "MEDIUM" if ps >= 3 else "LOW",
                "priority_score": ps,
                "max_buy": max_buy, "sweet_spot": sweet_spot,
                "median": a['median'], "floor": a['floor'],
                "format": "Reserve" if is_reserve else "NR",
                "signal": a['signal'],
                "risk_vg_pct": round(a['risk_nr'], 1) if a['risk_nr'] is not None else None,
                "volume": a['volume'], "sell_through": None,
                "trend": "—", "trend_pct": 0,
                "search_terms": search_terms,
                "condition_filter": ["Excellent", "Like New", "Very Good", "BNIB"],
                "papers_required": True,
                "action": "auto_evaluate",
                "notes": f"DJ config: {cfg_name}. {'Reserve account.' if is_reserve else 'NR safe.'}",
            })
    
    # ── Notable discoveries ──
    disc_targets = []
    for d in discoveries[:15]:
        a = d['analysis']
        if a['signal'] not in ('Strong', 'Normal'):
            continue
        if a['median'] < 2000 or a['median'] > 25000:
            continue
        disc_targets.append({
            "brand": d.get('brand', '?'), "reference": d.get('reference', '?'),
            "median": a['median'], "max_buy": a['max_buy_nr'],
            "signal": a['signal'], "volume": d.get('count', 0),
            "search_terms": [f"{d.get('brand','?')} {d.get('reference','?')}"],
            "action": "flag_for_review",
            "notes": "Not in core program. Flag only, do not auto-evaluate.",
        })
    
    targets.sort(key=lambda t: -t['priority_score'])
    
    # ── JSON brief → state/ ──
    brief = {
        "schema_version": 1,
        "generated_at": now.isoformat(),
        "source_report": source_file,
        "valid_until": "Next Grailzee Pro report (~2 weeks)",
        "sourcing_rules": {
            "us_inventory_only": True,
            "papers_required": True,
            "condition_minimum": "Very Good",
            "never_exceed_max_buy": True,
            "platform_priority": [
                {"platform": "facebook_groups", "type": "private dealer groups", "check_frequency": "daily"},
                {"platform": "whatsapp", "type": "dealer group chats", "check_frequency": "real_time"},
                {"platform": "ebay", "type": "BIN listings", "check_frequency": "twice_daily"},
                {"platform": "chrono24", "type": "US dealer listings", "check_frequency": "daily"},
                {"platform": "reddit", "type": "r/watchexchange", "check_frequency": "daily"},
            ],
            "keyword_filters": {
                "include": ["full set", "complete set", "box papers", "BNIB", "like new",
                            "excellent", "very good", "AD", "authorized"],
                "exclude": ["watch only", "no papers", "head only", "international",
                            "damaged", "for parts", "aftermarket", "rep", "homage"],
            },
        },
        "targets": targets,
        "discoveries": disc_targets,
        "summary": {
            "total_targets": len(targets),
            "high_priority": sum(1 for t in targets if t['priority'] == "HIGH"),
            "medium_priority": sum(1 for t in targets if t['priority'] == "MEDIUM"),
            "low_priority": sum(1 for t in targets if t['priority'] == "LOW"),
            "discoveries_flagged": len(disc_targets),
            "lowest_entry_point": min(t['sweet_spot'] for t in targets) if targets else 0,
            "highest_entry_point": max(t['max_buy'] for t in targets) if targets else 0,
        },
    }
    
    state_dir = os.path.join(os.path.dirname(output_dir), "state")
    os.makedirs(state_dir, exist_ok=True)
    json_path = os.path.join(state_dir, "sourcing_brief.json")
    with open(json_path, 'w') as f:
        json.dump(brief, f, indent=2, default=str)
    print(f"  Sourcing brief (JSON): {json_path}")
    
    # ── Markdown brief → output/ ──
    md = []
    md.append(f"# Vardalux Sourcing Brief — {now.strftime('%B %d, %Y')}")
    md.append(f"Source: {source_file}")
    md.append(f"Valid until next Grailzee Pro report (~2 weeks)\n")
    md.append(f"## Active Targets: {len(targets)} references\n")
    
    high = [t for t in targets if t['priority'] == 'HIGH']
    if high:
        md.append("### Priority: HIGH (hunt actively)\n")
        md.append("| Reference | Model | MAX BUY | Sweet Spot | Signal | Trend |")
        md.append("|-----------|-------|---------|------------|--------|-------|")
        for t in high:
            md.append(f"| {t['reference']} | {t['brand']} {t['model']} | "
                      f"**${t['max_buy']:,.0f}** | ${t['sweet_spot']:,.0f} | "
                      f"{t['signal']} | {t['trend']} |")
            if t['notes']:
                md.append(f"  *{t['notes']}*")
    
    med = [t for t in targets if t['priority'] == 'MEDIUM']
    if med:
        md.append("\n### Priority: MEDIUM (buy on sight if price works)\n")
        md.append("| Reference | Model | MAX BUY | Sweet Spot | Signal | Trend |")
        md.append("|-----------|-------|---------|------------|--------|-------|")
        for t in med:
            md.append(f"| {t['reference']} | {t['brand']} {t['model']} | "
                      f"**${t['max_buy']:,.0f}** | ${t['sweet_spot']:,.0f} | "
                      f"{t['signal']} | {t['trend']} |")
    
    low = [t for t in targets if t['priority'] == 'LOW']
    if low:
        md.append("\n### Priority: LOW (opportunistic only)\n")
        md.append("| Reference | Model | MAX BUY | Signal | Notes |")
        md.append("|-----------|-------|---------|--------|-------|")
        for t in low:
            md.append(f"| {t['reference']} | {t['brand']} {t['model']} | "
                      f"**${t['max_buy']:,.0f}** | {t['signal']} | {t['notes']} |")
    
    if disc_targets:
        md.append(f"\n### Watching: {len(disc_targets)} discovered references (flag only)\n")
        md.append("| Reference | Brand | Median | MAX BUY | Signal | Sales |")
        md.append("|-----------|-------|--------|---------|--------|-------|")
        for d in disc_targets:
            md.append(f"| {d['reference']} | {d['brand']} | "
                      f"${d['median']:,.0f} | ${d['max_buy']:,.0f} | "
                      f"{d['signal']} | {d['volume']} |")
    
    md.append("\n## Search Keywords\n")
    md.append("### Include (any match)")
    md.append("full set, complete set, box papers, BNIB, like new, excellent, very good, AD, authorized\n")
    md.append("### Exclude (skip listing)")
    md.append("watch only, no papers, head only, international, damaged, for parts, aftermarket, rep, homage\n")
    md.append("## Platform Scan Order\n")
    md.append("1. Private Facebook groups (daily)")
    md.append("2. WhatsApp dealer chats (real-time)")
    md.append("3. eBay BIN listings (twice daily)")
    md.append("4. Chrono24 US dealers (daily)")
    md.append("5. Reddit r/watchexchange (daily)\n")
    md.append("US inventory only. Never exceed MAX BUY. Papers required on every deal.\n")
    md.append(f"---\n*Generated {now.strftime('%B %d, %Y')} from {source_file}*")
    
    md_path = os.path.join(output_dir,
                           f"Vardalux_Sourcing_Brief_{now.strftime('%B%Y')}.md")
    with open(md_path, 'w') as f:
        f.write('\n'.join(md))
    print(f"  Sourcing brief (MD): {md_path}")
    
    return json_path, md_path


# ═══ MAIN ═══
def generate_output(folder, output_path=None):
    print("="*70); print("VARDALUX GRAILZEE ANALYZER"); print("="*70)
    reports = find_reports(folder)
    if not reports:
        print(f"ERROR: No reports found in {folder}"); return None
    print(f"\nFound {len(reports)} report(s) (sorted by date):")
    for r in reports: print(f"  {os.path.basename(r)}")

    current_file = reports[-1]
    print(f"\n--- Current: {os.path.basename(current_file)} ---")
    curr_sales, curr_st = parse_report(current_file)
    if not curr_sales: print("ERROR: No sales"); return None
    print(f"  Sales: {len(curr_sales)}")

    results, matched, discoveries = match_core_sales(curr_sales, curr_st)
    print(f"  Core matched: {len([k for k in results if not k.startswith('_')])}")
    if discoveries:
        print(f"  Discovered: {len(discoveries)}")
        for d in discoveries[:10]:
            print(f"    {d['brand']} {d['reference']} — {d['count']} sales, ${d['analysis']['median']:,.0f}")

    trends = []
    if len(reports) >= 2:
        prev_file = reports[-2]
        print(f"\n--- Previous: {os.path.basename(prev_file)} ---")
        prev_sales, prev_st = parse_report(prev_file)
        if prev_sales:
            prev_results, _, _ = match_core_sales(prev_sales, prev_st)
            trends = compare_periods(results, prev_results)
            notable = [t for t in trends if t['signals']]
            print(f"  Trends: {len(trends)}, Notable: {len(notable)}")
            for t in notable: print(f"    {t['model']}: {t['signal_str']}")

    if output_path is None:
        GDRIVE_OUTPUT = "/Users/ranbirchawla/Library/CloudStorage/GoogleDrive-ranbir.chawla@rnvillc.com/Shared drives/Vardalux Shared Drive/GrailzeeData/output"
        CHAT_OUTPUT = "/mnt/user-data/outputs"
        out_dir = GDRIVE_OUTPUT if os.path.isdir(GDRIVE_OUTPUT) else CHAT_OUTPUT
        os.makedirs(out_dir, exist_ok=True)
        output_path = os.path.join(out_dir,
                                    f"Vardalux_Grailzee_Buy_Targets_{datetime.now().strftime('%B%Y')}.xlsx")
    build_spreadsheet(results, trends, discoveries, output_path)
    build_summary(results, trends, discoveries,
                  os.path.basename(current_file), len(curr_sales),
                  os.path.dirname(output_path))
    build_sourcing_brief(results, trends, discoveries,
                         os.path.basename(current_file),
                         os.path.dirname(output_path))

    # Write analysis cache for OpenClaw deal evaluator
    if CACHE_AVAILABLE:
        try:
            save_analysis_cache(results, trends, discoveries,
                               os.path.basename(current_file))
            print("  Analysis cache updated for deal evaluator.")
        except Exception as e:
            print(f"  WARNING: Cache write failed: {e}")
            print("  Spreadsheet is fine. Deal evaluator may use stale data.")

    # Also copy spreadsheet and summary to persistent GrailzeeData folder
    persistent_output = "/Users/ranbirchawla/Library/CloudStorage/GoogleDrive-ranbir.chawla@rnvillc.com/Shared drives/Vardalux Shared Drive/GrailzeeData/output/"
    if os.path.isdir(persistent_output) and not output_path.startswith(persistent_output):
        import shutil
        for src in [output_path,
                    os.path.join(os.path.dirname(output_path),
                                 f"Vardalux_Grailzee_Analysis_{datetime.now().strftime('%B%Y')}.md"),
                    os.path.join(os.path.dirname(output_path),
                                 f"Vardalux_Sourcing_Brief_{datetime.now().strftime('%B%Y')}.md")]:
            if os.path.exists(src):
                try:
                    shutil.copy2(src, os.path.join(persistent_output, os.path.basename(src)))
                    print(f"  Persistent copy: {os.path.basename(src)}")
                except Exception as e:
                    print(f"  WARNING: Persistent copy failed for {os.path.basename(src)}: {e}")

    return {'current_file':os.path.basename(current_file),'total_sales':len(curr_sales),
            'results':results,'trends':trends,'discoveries':discoveries,'output_path':output_path}

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 analyze_report.py <folder> [--output <path>]"); sys.exit(1)
    folder = sys.argv[1]
    output = None
    if '--output' in sys.argv:
        idx = sys.argv.index('--output')
        if idx+1 < len(sys.argv): output = sys.argv[idx+1]
    generate_output(folder, output)

