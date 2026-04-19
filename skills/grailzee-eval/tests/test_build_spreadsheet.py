"""Tests for scripts.build_spreadsheet; branded openpyxl output.

Extraction from v1 analyze_report.py:383-598. v1/v2 structural equivalence:
sheet names and header columns match v1 (minus Raw Data tab, dropped in v2).

Test approach: structural assertions (sheet count, names, header row,
brand color hex, gold fill on MAX BUY, computed cell values). Not
byte-for-byte file comparison.

Fixture reference:
  79830RB: median=3200, st_pct=0.6, max_buy_nr=2910, risk_nr=8.0,
           signal=Strong, recommend_reserve=False, floor=3000, volume=5
"""

from pathlib import Path

import openpyxl
import pytest

from scripts.build_spreadsheet import build_spreadsheet, RICH_BLACK, WARM_GOLD


def _ref(brand: str = "Tudor", model: str = "BB GMT", reference: str = "79830RB",
         median: float = 3200, st_pct: float = 0.6, max_buy_nr: float = 2910,
         max_buy_res: float = 2770, risk_nr: float = 8.0, signal: str = "Strong",
         recommend_reserve: bool = False, floor: float = 3000, volume: int = 5,
         named: bool = True) -> dict:
    return {
        "brand": brand, "model": model, "reference": reference,
        "median": median, "st_pct": st_pct, "max_buy_nr": max_buy_nr,
        "max_buy_res": max_buy_res, "risk_nr": risk_nr, "signal": signal,
        "recommend_reserve": recommend_reserve, "floor": floor, "volume": volume,
        "named": named,
    }


def _empty_inputs() -> tuple:
    return ({}, {}, {}, {}, {}, {}, {})


def _fixture_inputs() -> tuple:
    all_results = {
        "references": {"79830RB": _ref()},
        "dj_configs": {},
        "unnamed": [],
    }
    trends = {"trends": [], "momentum": {}, "period_count": 1, "note": "Single report"}
    changes = {"emerged": [], "shifted": {}, "faded": [], "unnamed": []}
    breakouts = {"breakouts": [], "count": 0}
    watchlist = {"watchlist": [], "count": 0}
    brands_data = {"brands": {}, "count": 0}
    ledger_stats = {"trades": [], "summary": {"total_trades": 0}}
    return (all_results, trends, changes, breakouts, watchlist, brands_data, ledger_stats)


# ═══════════════════════════════════════════════════════════════════════
# Sheet structure
# ═══════════════════════════════════════════════════════════════════════


class TestSheetStructure:
    def test_sheet_count_no_trends(self, tmp_path):
        """No trends -> 2 sheets (Buy Targets + Quick Reference)."""
        inputs = _fixture_inputs()
        path = build_spreadsheet(*inputs, str(tmp_path))
        wb = openpyxl.load_workbook(path)
        assert len(wb.sheetnames) == 2
        assert wb.sheetnames == ["Buy Targets", "Quick Reference"]

    def test_sheet_count_with_trends(self, tmp_path):
        """With trends -> 3 sheets."""
        all_results, trends, changes, breakouts, watchlist, brands_data, ledger_stats = _fixture_inputs()
        trends["trends"] = [{
            "reference": "79830RB", "brand": "Tudor", "model": "BB GMT",
            "prev_median": 3000, "curr_median": 3200, "med_change": 200,
            "med_pct": 6.67, "prev_st": 0.5, "curr_st": 0.6,
            "st_change": 10.0, "prev_vol": 4, "curr_vol": 5,
            "signals": ["Momentum"], "signal_str": "Momentum",
        }]
        path = build_spreadsheet(all_results, trends, changes, breakouts,
                                 watchlist, brands_data, ledger_stats, str(tmp_path))
        wb = openpyxl.load_workbook(path)
        assert len(wb.sheetnames) == 3
        assert "Trends" in wb.sheetnames

    def test_header_row_contents(self, tmp_path):
        """Header row (row 5) matches expected columns."""
        inputs = _fixture_inputs()
        path = build_spreadsheet(*inputs, str(tmp_path))
        wb = openpyxl.load_workbook(path)
        ws = wb["Buy Targets"]
        headers = [ws.cell(row=5, column=i).value for i in range(1, 13)]
        expected = ["Brand", "Model", "Reference", "Median", "ST%",
                    "MAX BUY", "Risk(VG+)", "Signal", "Format", "Floor", "Vol", "Notes"]
        assert headers == expected


# ═══════════════════════════════════════════════════════════════════════
# Brand colors and styling
# ═══════════════════════════════════════════════════════════════════════


class TestBrandColors:
    def test_title_fill_rich_black(self, tmp_path):
        """Title cell A1 has RICH_BLACK fill."""
        inputs = _fixture_inputs()
        path = build_spreadsheet(*inputs, str(tmp_path))
        wb = openpyxl.load_workbook(path)
        ws = wb["Buy Targets"]
        fill = ws["A1"].fill
        assert fill.fgColor.rgb is not None
        assert fill.fgColor.rgb.endswith(RICH_BLACK)

    def test_max_buy_gold_fill(self, tmp_path):
        """MAX BUY column (col 6) on data row has WARM_GOLD fill."""
        inputs = _fixture_inputs()
        path = build_spreadsheet(*inputs, str(tmp_path))
        wb = openpyxl.load_workbook(path)
        ws = wb["Buy Targets"]
        # Data row is row 6 (row 4=section, row 5=headers, row 6=first data)
        cell = ws.cell(row=6, column=6)
        assert cell.fill.fgColor.rgb is not None
        assert cell.fill.fgColor.rgb.endswith(WARM_GOLD)


# ═══════════════════════════════════════════════════════════════════════
# Computed values
# ═══════════════════════════════════════════════════════════════════════


class TestComputedValues:
    def test_reference_row_values(self, tmp_path):
        """First data row contains expected reference values."""
        inputs = _fixture_inputs()
        path = build_spreadsheet(*inputs, str(tmp_path))
        wb = openpyxl.load_workbook(path)
        ws = wb["Buy Targets"]
        row = 6  # first data row
        assert ws.cell(row=row, column=1).value == "Tudor"
        assert ws.cell(row=row, column=2).value == "BB GMT"
        assert ws.cell(row=row, column=3).value == "79830RB"
        assert ws.cell(row=row, column=4).value == 3200  # median
        assert ws.cell(row=row, column=6).value == 2910  # max_buy_nr
        assert ws.cell(row=row, column=8).value == "Strong"
        assert ws.cell(row=row, column=9).value == "NR"
        assert ws.cell(row=row, column=11).value == 5  # volume

    def test_reserve_reference(self, tmp_path):
        """Reserve reference shows max_buy_res and Format=Reserve."""
        all_results = {
            "references": {"A17320": _ref(
                brand="Breitling", model="SO Heritage", reference="A17320",
                median=2400, max_buy_nr=2140, max_buy_res=2100,
                risk_nr=45.0, signal="Careful", recommend_reserve=True,
                floor=2100, volume=3,
            )},
            "dj_configs": {},
        }
        path = build_spreadsheet(all_results, {}, {}, {}, {}, {}, {}, str(tmp_path))
        wb = openpyxl.load_workbook(path)
        ws = wb["Buy Targets"]
        row = 6
        assert ws.cell(row=row, column=6).value == 2100  # max_buy_res
        assert ws.cell(row=row, column=9).value == "Reserve"


# ═══════════════════════════════════════════════════════════════════════
# DJ Configs section
# ═══════════════════════════════════════════════════════════════════════


class TestDJConfigs:
    def test_dj_config_section_present(self, tmp_path):
        """DJ configs appear in spreadsheet when present."""
        all_results = {
            "references": {"79830RB": _ref()},
            "dj_configs": {"Black/Oyster": _ref(
                brand="Rolex", model="DJ 41 Black/Oyster", reference="126300",
                median=9500, max_buy_nr=8900,
            )},
        }
        path = build_spreadsheet(all_results, {}, {}, {}, {}, {}, {}, str(tmp_path))
        wb = openpyxl.load_workbook(path)
        ws = wb["Buy Targets"]
        # Search for DJ section header
        found = False
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and "DATEJUST 126300" in str(val):
                found = True
                break
        assert found


# ═══════════════════════════════════════════════════════════════════════
# Empty input
# ═══════════════════════════════════════════════════════════════════════


class TestEmptyInput:
    def test_empty_produces_valid_xlsx(self, tmp_path):
        """Empty all_results -> valid xlsx with headers, no data rows."""
        path = build_spreadsheet(*_empty_inputs(), str(tmp_path))
        assert Path(path).exists()
        wb = openpyxl.load_workbook(path)
        assert "Buy Targets" in wb.sheetnames

    def test_output_filename(self, tmp_path):
        """Output filename follows convention."""
        path = build_spreadsheet(*_empty_inputs(), str(tmp_path))
        assert "Vardalux_Grailzee_Buy_Targets_" in Path(path).name
        assert path.endswith(".xlsx")


# ═══════════════════════════════════════════════════════════════════════
# v1/v2 structural equivalence
# ═══════════════════════════════════════════════════════════════════════


class TestV1V2Equivalence:
    def test_header_columns_match_v1(self, tmp_path):
        """v2 header columns match v1 exactly."""
        inputs = _fixture_inputs()
        path = build_spreadsheet(*inputs, str(tmp_path))
        wb = openpyxl.load_workbook(path)
        ws = wb["Buy Targets"]
        headers = [ws.cell(row=5, column=i).value for i in range(1, 13)]
        v1_headers = ["Brand", "Model", "Reference", "Median", "ST%",
                      "MAX BUY", "Risk(VG+)", "Signal", "Format", "Floor", "Vol", "Notes"]
        assert headers == v1_headers

    def test_quick_reference_present(self, tmp_path):
        """Quick Reference tab present (matches v1)."""
        inputs = _fixture_inputs()
        path = build_spreadsheet(*inputs, str(tmp_path))
        wb = openpyxl.load_workbook(path)
        assert "Quick Reference" in wb.sheetnames
