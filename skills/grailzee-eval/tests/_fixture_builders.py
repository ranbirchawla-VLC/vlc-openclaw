"""Programmatic builders for synthetic Grailzee Pro Excel fixtures.

Avoids committing binary .xlsx files to git. Each test that needs an
Excel fixture calls a builder function, which writes a fresh workbook
to a tmp_path location.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl


def build_minimal_report(
    path: str | Path,
    sales_rows: list[dict] | None = None,
    include_top_selling: bool = True,
    sell_through_format: str = "percent_string",
    extra_auctions_column: str | None = None,
    omit_auctions_sheet: bool = False,
    omit_required_column: str | None = None,
    header_style: str = "w1",
) -> Path:
    """Write a synthetic Grailzee Pro report to `path`.

    Args:
        path: output .xlsx path
        sales_rows: list of dicts with keys matching Auctions Sold columns.
            If None, uses 3 default rows.
        include_top_selling: whether to add Top Selling Watches sheet
        sell_through_format: "percent_string" ("23%"), "fraction" (0.23),
            or "integer_percent" (23)
        extra_auctions_column: if set, adds an extra column with this name
        omit_auctions_sheet: if True, skips creating Auctions Sold sheet
        omit_required_column: if set, omits this column from Auctions Sold header
    """
    path = Path(path)
    wb = openpyxl.Workbook()

    if sales_rows is None:
        sales_rows = DEFAULT_SALES_ROWS

    # ─── Auctions Sold ───────────────────────────────────────────
    if not omit_auctions_sheet:
        ws = wb.active
        ws.title = "Auctions Sold"

        if header_style == "w2":
            sold_at_header = "Sold At"
            dial_header = "Dial Numbers"
        else:
            sold_at_header = "Sold at"
            dial_header = "Dial"
        headers = [
            sold_at_header, "Auction", "Make", "Model", "Reference Number",
            "Sold For", "Condition", "Year", "Papers", "Box",
            dial_header, "URL",
        ]
        if extra_auctions_column:
            headers.append(extra_auctions_column)
        if omit_required_column and omit_required_column in headers:
            headers.remove(omit_required_column)

        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)

        for ri, row in enumerate(sales_rows, 2):
            for ci, h in enumerate(headers, 1):
                key = h.lower().replace(" ", "_")
                # Map header names to row dict keys
                key_map = {
                    "sold_at": "sold_at",
                    "auction": "title",
                    "make": "make",
                    "model": "model",
                    "reference_number": "reference",
                    "sold_for": "sold_for",
                    "condition": "condition",
                    "year": "year",
                    "papers": "papers",
                    "box": "box",
                    "dial": "dial",
                    "dial_numbers": "dial",
                    "url": "url",
                }
                mapped = key_map.get(key, key)
                val = row.get(mapped)
                if val is not None:
                    ws.cell(row=ri, column=ci, value=val)
    else:
        ws = wb.active
        ws.title = "Sheet1"

    # ─── Top Selling Watches ─────────────────────────────────────
    if include_top_selling:
        ws2 = wb.create_sheet("Top Selling Watches")
        ts_headers = [
            "Watch", "Reference Number", "Sales", "Total",
            "Sell-Through Rate", "Average Price",
        ]
        for ci, h in enumerate(ts_headers, 1):
            ws2.cell(row=1, column=ci, value=h)

        # Aggregate sales_rows by reference for top selling
        ref_agg: dict[str, dict] = {}
        for row in sales_rows:
            ref = row.get("reference")
            if ref is None:
                continue
            ref_str = str(ref)
            if ref_str.endswith(".0"):
                ref_str = ref_str[:-2]
            if ref_str not in ref_agg:
                ref_agg[ref_str] = {
                    "model": row.get("model", ""),
                    "reference": row.get("reference"),
                    "sales": 0,
                    "total": 0,
                    "sum_price": 0,
                }
            ref_agg[ref_str]["sales"] += 1
            ref_agg[ref_str]["total"] += 1
            price_val = row.get("sold_for") or 0
            if isinstance(price_val, str):
                try:
                    price_val = float(price_val.replace("$", "").replace(",", ""))
                except ValueError:
                    price_val = 0
            ref_agg[ref_str]["sum_price"] += price_val

        for ri, (ref_str, agg) in enumerate(ref_agg.items(), 2):
            ws2.cell(row=ri, column=1, value=agg["model"])
            ws2.cell(row=ri, column=2, value=agg["reference"])
            ws2.cell(row=ri, column=3, value=agg["sales"])
            ws2.cell(row=ri, column=4, value=max(agg["total"], agg["sales"] + 2))
            st_val = agg["sales"] / max(agg["total"], agg["sales"] + 2)
            if sell_through_format == "percent_string":
                ws2.cell(row=ri, column=5, value=f"{st_val * 100:.0f}%")
            elif sell_through_format == "fraction":
                ws2.cell(row=ri, column=5, value=round(st_val, 2))
            elif sell_through_format == "integer_percent":
                ws2.cell(row=ri, column=5, value=int(st_val * 100))
            avg_price = agg["sum_price"] / agg["sales"] if agg["sales"] else 0
            ws2.cell(row=ri, column=6, value=round(avg_price))

    wb.save(str(path))
    return path


# Default 3 sales rows for minimal fixture
DEFAULT_SALES_ROWS = [
    {
        "sold_at": datetime(2026, 2, 5, 14, 30, 0),
        "title": "No Reserve - 2020 Tudor Black Bay GMT Pepsi 79830RB",
        "make": "Tudor",
        "model": "Tudor Black Bay GMT",
        "reference": "79830RB",
        "sold_for": 3200,
        "condition": "Very Good",
        "year": 2020,
        "papers": "Yes",
        "box": "Yes",
        "dial": "No Numerals",
        "url": "https://grailzee.com/products/tudor-black-bay-gmt-pepsi-79830rb-1",
    },
    {
        "sold_at": datetime(2026, 2, 7, 10, 0, 0),
        "title": "No Reserve - 2021 Breitling Superocean Heritage A17320",
        "make": "Breitling",
        "model": "Breitling Superocean Heritage 42",
        "reference": "A17320",
        "sold_for": 2400,
        "condition": "Excellent",
        "year": 2021,
        "papers": "Yes",
        "box": "No",
        "dial": "Arabic Numerals",
        "url": "https://grailzee.com/products/breitling-superocean-heritage-a17320-1",
    },
    {
        "sold_at": datetime(2026, 2, 9, 16, 45, 0),
        "title": "No Reserve - 2019 Tudor Black Bay GMT Pepsi 79830RB Full Set",
        "make": "Tudor",
        "model": "Tudor Black Bay GMT",
        "reference": "79830RB",
        "sold_for": 3050,
        "condition": "Like New",
        "year": 2019,
        "papers": "Yes",
        "box": "Yes",
        "dial": "No Numerals",
        "url": "https://grailzee.com/products/tudor-black-bay-gmt-pepsi-79830rb-full-set-2",
    },
]
