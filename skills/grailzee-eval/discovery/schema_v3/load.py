"""Load April W1 and W2 xlsx reports into in-memory row dicts.

Phase 1 discovery-only. Deleted after Phase 2 ship.
Reads headers from row 1, rows from row 2+. Preserves raw cell values
(no canonicalization here; that is what Phase 1 is investigating).
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

DRIVE = Path(
    "/Users/ranbirchawla/Library/CloudStorage/"
    "GoogleDrive-ranbir.chawla@rnvillc.com/Shared drives/"
    "Vardalux Shared Drive/GrailzeeData"
)
W1_PATH = DRIVE / "reports" / "Grailzee Pro Bi-Weekly Report - April W1.xlsx"
W2_PATH = DRIVE / "reports" / "Grailzee Pro Bi-Weekly Report - April W2.xlsx"


@dataclass
class Report:
    label: str
    path: Path
    sheet_name: str
    headers: list[str]
    rows: list[dict[str, Any]]
    aggregate_sheets: dict[str, list[list[Any]]]


def load(path: Path, label: str) -> Report:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    primary_sheet = wb.sheetnames[0]
    ws = wb[primary_sheet]
    row_iter = ws.iter_rows(values_only=True)
    headers_raw = next(row_iter)
    headers = [h if h is not None else f"_col{i}" for i, h in enumerate(headers_raw)]
    rows: list[dict[str, Any]] = []
    for r in row_iter:
        if r is None or all(c is None for c in r):
            continue
        rows.append({headers[i]: r[i] if i < len(r) else None for i in range(len(headers))})
    aggregate_sheets: dict[str, list[list[Any]]] = {}
    for name in wb.sheetnames[1:]:
        ws2 = wb[name]
        aggregate_sheets[name] = [list(r) for r in ws2.iter_rows(values_only=True)]
    return Report(
        label=label,
        path=path,
        sheet_name=primary_sheet,
        headers=headers,
        rows=rows,
        aggregate_sheets=aggregate_sheets,
    )


def load_both() -> tuple[Report, Report]:
    return load(W1_PATH, "W1"), load(W2_PATH, "W2")


if __name__ == "__main__":
    w1, w2 = load_both()
    for r in (w1, w2):
        print(f"=== {r.label}: {r.path.name}")
        print(f"  primary sheet: {r.sheet_name}")
        print(f"  headers ({len(r.headers)}): {r.headers}")
        print(f"  rows: {len(r.rows)}")
        print(f"  aggregate sheets: {list(r.aggregate_sheets.keys())}")
