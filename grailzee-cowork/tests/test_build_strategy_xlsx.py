"""Tests for build_strategy_xlsx — the Phase 24b archival XLSX builder.

Validates:
- Minimal payload (cycle_focus only) produces exactly 2 sheets
- Each populated decision section adds its own sheet; null sections omit it
- Sheet order is deterministic: Cycle Summary → Targets → Monthly Goals
  → Quarterly → Config Updates
- Cycle Summary header carries the cycle_id and session_mode
- target_margin_fraction renders as a percent string ("5.0%"), not a
  numeric fraction
- Vardalux brand colors are actually applied (rich_black / warm_gold /
  deep_teal)
- Config Updates table reports Yes/No per sub-config and surfaces
  envelope notes
- Output file is openable by openpyxl round-trip
- Parent directory is created on demand
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from _fixtures import (
    FAKE_CYCLE_ID,
    make_strategy_output,
)

from grailzee_bundle.build_strategy_xlsx import (
    VARDALUX_COLORS,
    build_strategy_xlsx,
)


# ─── sheet presence / order ──────────────────────────────────────────

def test_minimal_payload_yields_cycle_summary_and_targets(tmp_path: Path) -> None:
    """Default fixture has cycle_focus only → 2 sheets."""
    payload = make_strategy_output(cycle_id=FAKE_CYCLE_ID)
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")

    wb = load_workbook(out)
    assert wb.sheetnames == ["Cycle Summary", "Targets"]


def test_all_four_decisions_yields_five_sheets_in_order(tmp_path: Path) -> None:
    payload = make_strategy_output(
        cycle_id=FAKE_CYCLE_ID,
        include_monthly=True,
        include_quarterly=True,
        include_configs=("signal_thresholds", "margin_config"),
    )
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")

    wb = load_workbook(out)
    assert wb.sheetnames == [
        "Cycle Summary",
        "Targets",
        "Monthly Goals",
        "Quarterly",
        "Config Updates",
    ]


def test_cycle_focus_null_omits_targets_sheet(tmp_path: Path) -> None:
    """A payload whose only decision is monthly_goals gets Cycle Summary
    (always) + Monthly Goals — no Targets sheet, since cycle_focus is null."""
    payload = make_strategy_output(
        cycle_id=FAKE_CYCLE_ID,
        include_cycle_focus=False,
        include_monthly=True,
        session_mode="monthly_review",
    )
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")

    wb = load_workbook(out)
    assert wb.sheetnames == ["Cycle Summary", "Monthly Goals"]


def test_only_quarterly_produces_cycle_summary_and_quarterly(tmp_path: Path) -> None:
    payload = make_strategy_output(
        cycle_id=FAKE_CYCLE_ID,
        include_cycle_focus=False,
        include_quarterly=True,
        session_mode="quarterly_allocation",
    )
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")
    assert load_workbook(out).sheetnames == ["Cycle Summary", "Quarterly"]


# ─── cycle summary content ───────────────────────────────────────────

def test_cycle_summary_title_carries_cycle_id(tmp_path: Path) -> None:
    payload = make_strategy_output(cycle_id="cycle_2026-07")
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")

    ws = load_workbook(out)["Cycle Summary"]
    assert ws["A1"].value == "Vardalux Grailzee — cycle_2026-07"


def test_cycle_summary_surfaces_session_mode_and_produced_by(tmp_path: Path) -> None:
    payload = make_strategy_output(
        cycle_id=FAKE_CYCLE_ID,
        session_mode="config_tuning",
        include_cycle_focus=False,
        include_configs=("signal_thresholds",),
        produced_by="grailzee-strategy/0.1.0",
    )
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")

    ws = load_workbook(out)["Cycle Summary"]
    assert ws["B3"].value == "config_tuning"
    assert ws["B5"].value == "grailzee-strategy/0.1.0"


def test_target_margin_fraction_renders_as_percent_string(tmp_path: Path) -> None:
    """5.0% must appear as a formatted string, not the raw 0.05 float —
    operators read this directly."""
    payload = make_strategy_output(cycle_id=FAKE_CYCLE_ID)
    # Default fraction is 0.05. Sanity check that's what the fixture gives us.
    assert payload["decisions"]["cycle_focus"]["target_margin_fraction"] == 0.05
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")

    ws = load_workbook(out)["Cycle Summary"]
    assert ws["B10"].value == "5.0%"


def test_target_margin_fraction_custom_value(tmp_path: Path) -> None:
    payload = make_strategy_output(cycle_id=FAKE_CYCLE_ID)
    payload["decisions"]["cycle_focus"]["target_margin_fraction"] = 0.125
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")

    ws = load_workbook(out)["Cycle Summary"]
    assert ws["B10"].value == "12.5%"


# ─── style application ───────────────────────────────────────────────

def test_title_cell_uses_rich_black_fill(tmp_path: Path) -> None:
    """Regression guard: brand palette is actually wired to the workbook,
    not silently dropped."""
    payload = make_strategy_output(cycle_id=FAKE_CYCLE_ID)
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")

    ws = load_workbook(out)["Cycle Summary"]
    title = ws["A1"]
    # openpyxl prefixes 8-char ARGB on load; the last 6 chars are the RGB.
    rgb = (title.fill.fgColor.rgb or "")[-6:].upper()
    assert rgb == VARDALUX_COLORS["rich_black"].upper()
    assert title.font.bold is True


# ─── config updates table ────────────────────────────────────────────

def test_config_updates_table_reflects_yes_no_per_sub(tmp_path: Path) -> None:
    payload = make_strategy_output(
        cycle_id=FAKE_CYCLE_ID,
        include_cycle_focus=False,
        include_configs=("signal_thresholds", "margin_config"),
        session_mode="config_tuning",
    )
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")

    ws = load_workbook(out)["Config Updates"]
    # Header at row 6. Rows 7-12 are the six sub-configs in fixed order.
    rows = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(7, 13)}
    assert rows["signal_thresholds"] == "Yes"
    assert rows["margin_config"] == "Yes"
    assert rows["scoring_thresholds"] == "No"
    assert rows["momentum_thresholds"] == "No"
    assert rows["window_config"] == "No"
    assert rows["premium_config"] == "No"


def test_config_updates_surfaces_envelope_notes(tmp_path: Path) -> None:
    payload = make_strategy_output(
        cycle_id=FAKE_CYCLE_ID,
        include_cycle_focus=False,
        include_configs=("signal_thresholds",),
        session_mode="config_tuning",
    )
    # Populated sub should carry the template "Session retune." note.
    payload["decisions"]["config_updates"]["signal_thresholds"]["notes"] = "Tightened Tudor band"
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")

    ws = load_workbook(out)["Config Updates"]
    notes_for_signal = next(
        ws.cell(row=r, column=3).value
        for r in range(7, 13)
        if ws.cell(row=r, column=1).value == "signal_thresholds"
    )
    assert notes_for_signal == "Tightened Tudor band"


# ─── targets sheet ───────────────────────────────────────────────────

def test_targets_sheet_renders_override_dash_when_null(tmp_path: Path) -> None:
    payload = make_strategy_output(cycle_id=FAKE_CYCLE_ID)
    out = build_strategy_xlsx(payload, tmp_path / "brief.xlsx")

    ws = load_workbook(out)["Targets"]
    # Row 1 is headers, row 2 is the first (only) target from the fixture.
    assert ws.cell(row=2, column=1).value == "79830RB"
    assert ws.cell(row=2, column=5).value == "—"


# ─── file handling ───────────────────────────────────────────────────

def test_parent_directory_created_on_demand(tmp_path: Path) -> None:
    payload = make_strategy_output(cycle_id=FAKE_CYCLE_ID)
    nested = tmp_path / "output" / "briefs" / "nested" / "brief.xlsx"
    assert not nested.parent.exists()

    out = build_strategy_xlsx(payload, nested)
    assert out == nested
    assert nested.exists()


def test_returns_path_and_file_is_readable(tmp_path: Path) -> None:
    payload = make_strategy_output(cycle_id=FAKE_CYCLE_ID)
    target = tmp_path / "brief.xlsx"
    out = build_strategy_xlsx(payload, target)

    assert out == target
    assert target.exists()
    # Round-trip parse must succeed — guards against "wrote a truncated file"
    # style bugs that would otherwise only surface when the operator opens it.
    wb = load_workbook(target)
    assert "Cycle Summary" in wb.sheetnames
