"""Tests for the Phase 24b archive-write leg of apply_strategy_output.

The archive is a best-effort operator artifact — three files (JSON, XLSX,
MD) deposited into ``<grailzee_root>/output/briefs/``. State writes are
authoritative and must never be rolled back by an archive failure.

Coverage:
- Happy path: all three files created with ``<cycle_id>_...`` names
- Archive JSON round-trips byte-equal to the validated payload
- Archive MD matches session_artifacts.cycle_brief_md
- Archive XLSX opens via openpyxl round-trip
- archive_errors empty on success
- One file-level failure is captured without blocking the other two
- Directory-level failure is captured and no partial writes land
- write_archive=False disables the leg (state still writes)
- apply_strategy_output returns the archive fields alongside roles_written
- main() CLI surfaces archive_files_written + archive_errors in summary
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from _fixtures import (
    FAKE_CYCLE_ID,
    build_fake_grailzee_tree,
    make_strategy_output,
    write_strategy_output,
)

from grailzee_bundle.unpack_bundle import (
    _write_strategy_archive,
    apply_strategy_output,
    main,
)


# ─── _write_strategy_archive ─────────────────────────────────────────

def test_writes_all_three_files_with_cycle_id_prefix(tmp_path: Path) -> None:
    payload = make_strategy_output(cycle_id=FAKE_CYCLE_ID)
    files_written, errors = _write_strategy_archive(payload, tmp_path / "briefs")

    assert errors == []
    assert set(files_written) == {
        f"{FAKE_CYCLE_ID}_strategy_output.json",
        f"{FAKE_CYCLE_ID}_strategy_brief.xlsx",
        f"{FAKE_CYCLE_ID}_strategy_brief.md",
    }
    for name in files_written:
        assert (tmp_path / "briefs" / name).exists()


def test_archive_json_round_trips_to_payload(tmp_path: Path) -> None:
    """Auditability contract: operators must be able to reload the
    archived JSON and see exactly what was applied."""
    payload = make_strategy_output(
        cycle_id=FAKE_CYCLE_ID, include_monthly=True,
    )
    _write_strategy_archive(payload, tmp_path / "briefs")
    archived = json.loads(
        (tmp_path / "briefs" / f"{FAKE_CYCLE_ID}_strategy_output.json").read_text()
    )
    assert archived == payload


def test_archive_md_matches_session_artifacts(tmp_path: Path) -> None:
    payload = make_strategy_output(
        cycle_id=FAKE_CYCLE_ID,
        cycle_brief_md="# April Cycle\n\nFocus: Tudor GMT steel.\n",
    )
    _write_strategy_archive(payload, tmp_path / "briefs")
    md = (tmp_path / "briefs" / f"{FAKE_CYCLE_ID}_strategy_brief.md").read_text()
    assert md == "# April Cycle\n\nFocus: Tudor GMT steel.\n"


def test_archive_xlsx_readable_round_trip(tmp_path: Path) -> None:
    payload = make_strategy_output(
        cycle_id=FAKE_CYCLE_ID,
        include_monthly=True,
        include_quarterly=True,
    )
    _write_strategy_archive(payload, tmp_path / "briefs")
    wb = load_workbook(tmp_path / "briefs" / f"{FAKE_CYCLE_ID}_strategy_brief.xlsx")
    assert "Cycle Summary" in wb.sheetnames
    assert "Monthly Goals" in wb.sheetnames
    assert "Quarterly" in wb.sheetnames


def test_single_file_failure_captured_others_proceed(tmp_path: Path) -> None:
    """Create a DIRECTORY at the JSON file's target path so write_text
    fails with IsADirectoryError (an OSError subclass). The XLSX and MD
    writes must still succeed — best-effort semantics."""
    briefs = tmp_path / "briefs"
    briefs.mkdir()
    blocked_json = briefs / f"{FAKE_CYCLE_ID}_strategy_output.json"
    blocked_json.mkdir()  # now write_text to this path fails

    payload = make_strategy_output(cycle_id=FAKE_CYCLE_ID)
    files_written, errors = _write_strategy_archive(payload, briefs)

    assert set(files_written) == {
        f"{FAKE_CYCLE_ID}_strategy_brief.xlsx",
        f"{FAKE_CYCLE_ID}_strategy_brief.md",
    }
    assert len(errors) == 1
    assert errors[0]["file"] == f"{FAKE_CYCLE_ID}_strategy_output.json"
    assert errors[0]["error"]  # non-empty error message


def test_directory_creation_failure_short_circuits(tmp_path: Path) -> None:
    """If briefs_dir cannot be created (e.g. a FILE exists at that path),
    emit one directory-level error rather than three identical
    parent-missing errors."""
    blocker = tmp_path / "briefs"
    blocker.write_text("not a directory")

    payload = make_strategy_output(cycle_id=FAKE_CYCLE_ID)
    files_written, errors = _write_strategy_archive(payload, blocker)

    assert files_written == []
    assert len(errors) == 1
    assert errors[0]["file"] == str(blocker)


# ─── apply_strategy_output integration ───────────────────────────────

def test_apply_returns_archive_fields_on_success(tmp_path: Path) -> None:
    paths = build_fake_grailzee_tree(tmp_path)
    json_path = tmp_path / "strategy_output.json"
    write_strategy_output(json_path, make_strategy_output(cycle_id=FAKE_CYCLE_ID))

    result = apply_strategy_output(json_path, paths["root"])

    assert result["archive_errors"] == []
    assert set(result["archive_files_written"]) == {
        f"{FAKE_CYCLE_ID}_strategy_output.json",
        f"{FAKE_CYCLE_ID}_strategy_brief.xlsx",
        f"{FAKE_CYCLE_ID}_strategy_brief.md",
    }
    # Archive files land under the grailzee_root's briefs dir
    for name in result["archive_files_written"]:
        assert (paths["briefs"] / name).exists()


def test_apply_write_archive_false_skips_archive_leg(tmp_path: Path) -> None:
    paths = build_fake_grailzee_tree(tmp_path)
    json_path = tmp_path / "strategy_output.json"
    write_strategy_output(json_path, make_strategy_output(cycle_id=FAKE_CYCLE_ID))

    pre_briefs = set(p.name for p in paths["briefs"].iterdir())
    result = apply_strategy_output(json_path, paths["root"], write_archive=False)

    assert result["archive_files_written"] == []
    assert result["archive_errors"] == []
    # State still committed
    assert "cycle_focus" in result["roles_written"]
    # Briefs dir untouched — no new archive files
    post_briefs = set(p.name for p in paths["briefs"].iterdir())
    assert post_briefs == pre_briefs


def test_apply_archive_failure_does_not_block_state_commit(tmp_path: Path) -> None:
    """If archive writes fail, state writes must remain committed —
    archive is best-effort and must not roll back the primary apply."""
    paths = build_fake_grailzee_tree(tmp_path)
    # Sabotage the briefs dir: replace it with a regular file so mkdir
    # fails and every archive write is skipped. Clear any fixture
    # contents first since rmdir requires an empty directory.
    for p in paths["briefs"].iterdir():
        p.unlink()
    paths["briefs"].rmdir()
    paths["briefs"].write_text("blocker")

    json_path = tmp_path / "strategy_output.json"
    write_strategy_output(json_path, make_strategy_output(cycle_id=FAKE_CYCLE_ID))

    result = apply_strategy_output(json_path, paths["root"])

    # State write landed
    assert "cycle_focus" in result["roles_written"]
    focus = json.loads(paths["cycle_focus"].read_text())
    assert focus["targets"][0]["reference"] == "79830RB"
    # Archive failure surfaced, not raised
    assert result["archive_files_written"] == []
    assert len(result["archive_errors"]) == 1


# ─── main() CLI surface ──────────────────────────────────────────────

def test_main_json_path_includes_archive_fields(
    tmp_path: Path, capsys: pytest.CaptureFixture
) -> None:
    paths = build_fake_grailzee_tree(tmp_path)
    json_path = tmp_path / "strategy_output.json"
    write_strategy_output(json_path, make_strategy_output(cycle_id=FAKE_CYCLE_ID))

    rc = main([str(json_path), "--grailzee-root", str(paths["root"])])
    assert rc == 0
    summary = json.loads(capsys.readouterr().out)
    assert summary["archive_errors"] == []
    assert f"{FAKE_CYCLE_ID}_strategy_brief.xlsx" in summary["archive_files_written"]
    # payload still stripped from CLI output
    assert "payload" not in summary


def test_main_zip_path_does_not_emit_archive_fields(
    tmp_path: Path, capsys: pytest.CaptureFixture
) -> None:
    """The .zip (Phase 24a) path is unchanged — no archive leg, so no
    archive_files_written / archive_errors keys in its summary."""
    from _fixtures import build_inbound_bundle_zip
    paths = build_fake_grailzee_tree(tmp_path)
    zip_path = tmp_path / "inbound.zip"
    build_inbound_bundle_zip(zip_path, cycle_id=FAKE_CYCLE_ID)

    rc = main([str(zip_path), "--grailzee-root", str(paths["root"])])
    assert rc == 0
    summary = json.loads(capsys.readouterr().out)
    assert "archive_files_written" not in summary
    assert "archive_errors" not in summary
