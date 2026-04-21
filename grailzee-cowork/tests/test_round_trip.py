"""Round-trip integration tests: OUTBOUND → simulated Chat editing → INBOUND.

These tests exercise bundle-format compatibility end-to-end. They confirm
that the manifest schema is stable between the two sides, that the
OUTBOUND ``cycle_focus`` role (Phase A.5: previously ``cycle_focus_current``)
round-trips to the INBOUND ``cycle_focus`` role, and that validation +
atomic write commit the Chat session's edits into state/ without
corrupting unrelated files.
"""

from __future__ import annotations

import hashlib
import json
import zipfile
from pathlib import Path

import pytest

from openpyxl import load_workbook

from grailzee_bundle.build_bundle import build_outbound_bundle
from grailzee_bundle.unpack_bundle import apply_strategy_output, unpack_inbound_bundle
from _fixtures import (
    FAKE_CYCLE_ID,
    build_fake_grailzee_tree,
    make_strategy_output,
    write_strategy_output,
)


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _extract_role_bytes(zip_path: Path, role: str) -> bytes:
    with zipfile.ZipFile(zip_path, "r") as zf:
        manifest = json.loads(zf.read("manifest.json"))
        entry = next(f for f in manifest["files"] if f["role"] == role)
        return zf.read(entry["path"])


def _build_chat_inbound_zip(
    dest: Path,
    *,
    cycle_id: str,
    cycle_focus_bytes: bytes,
    monthly_goals_bytes: bytes,
    quarterly_allocation_bytes: bytes,
) -> Path:
    """Simulate what a Chat strategy session would hand back: the same three
    role files from the outbound bundle, possibly edited, rewrapped under
    the INBOUND manifest shape. A.5 aligned the OUTBOUND role name with
    INBOUND (both sides now speak ``cycle_focus``)."""
    payloads = [
        ("cycle_focus", "cycle_focus.json", cycle_focus_bytes),
        ("monthly_goals", "monthly_goals.json", monthly_goals_bytes),
        (
            "quarterly_allocation",
            "quarterly_allocation.json",
            quarterly_allocation_bytes,
        ),
    ]
    manifest = {
        "manifest_version": 1,
        "bundle_kind": "inbound",
        "generated_at": "2026-04-15T15:00:00Z",
        "cycle_id": cycle_id,
        "source": "chat-strategy-session",
        "scope": {"month_boundary": False, "quarter_boundary": False},
        "files": [
            {
                "path": arcname,
                "role": role,
                "sha256": _sha256(data),
                "size_bytes": len(data),
            }
            for role, arcname, data in payloads
        ],
    }
    with zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("manifest.json", json.dumps(manifest))
        for _role, arcname, data in payloads:
            zf.writestr(arcname, data)
    return dest


# ─── round-trip ─────────────────────────────────────────────────────


def test_unchanged_round_trip_is_byte_equivalent(tmp_path):
    """Build outbound → repackage the same three role files as inbound →
    unpack. State files end up byte-identical to their pre-outbound state."""
    paths = build_fake_grailzee_tree(tmp_path)
    before = {
        "cycle_focus": paths["cycle_focus"].read_bytes(),
        "monthly_goals": paths["monthly_goals"].read_bytes(),
        "quarterly_allocation": paths["quarterly_allocation"].read_bytes(),
    }

    outbound = build_outbound_bundle(tmp_path)

    # A.5: role renamed cycle_focus_current -> cycle_focus.
    cycle_focus_bytes = _extract_role_bytes(outbound, "cycle_focus")
    monthly_bytes = _extract_role_bytes(outbound, "monthly_goals")
    quarterly_bytes = _extract_role_bytes(outbound, "quarterly_allocation")

    inbound = _build_chat_inbound_zip(
        tmp_path / "chat_returned.zip",
        cycle_id=FAKE_CYCLE_ID,
        cycle_focus_bytes=cycle_focus_bytes,
        monthly_goals_bytes=monthly_bytes,
        quarterly_allocation_bytes=quarterly_bytes,
    )

    result = unpack_inbound_bundle(inbound, tmp_path)
    assert result["cycle_id"] == FAKE_CYCLE_ID
    assert result["roles_written"] == [
        "cycle_focus",
        "monthly_goals",
        "quarterly_allocation",
    ]

    assert paths["cycle_focus"].read_bytes() == before["cycle_focus"]
    assert paths["monthly_goals"].read_bytes() == before["monthly_goals"]
    assert paths["quarterly_allocation"].read_bytes() == before["quarterly_allocation"]


def test_edited_round_trip_commits_chat_edits(tmp_path):
    """Simulate Chat editing the three payload files; verify state/ reflects
    the edits and unrelated files (analysis_cache, trade_ledger, brief) are
    untouched."""
    paths = build_fake_grailzee_tree(tmp_path)
    cache_before = paths["cache"].read_bytes()
    ledger_before = paths["ledger"].read_bytes()
    brief_before = paths["brief"].read_bytes()

    outbound = build_outbound_bundle(tmp_path)

    # Pull the three outbound role payloads, edit each to represent a
    # strategy-session decision.
    cf = json.loads(_extract_role_bytes(outbound, "cycle_focus"))
    cf["focus_refs"] = ["79830RB", "210.30", "NEW_FROM_CHAT"]
    cf["chat_commentary"] = "Added SMD Ceramic blue dial based on Q2 momentum"

    mg = json.loads(_extract_role_bytes(outbound, "monthly_goals"))
    mg["revenue_target"] = 52000  # edited up
    mg["deal_count_target"] = 7

    qa = json.loads(_extract_role_bytes(outbound, "quarterly_allocation"))
    qa["allocations"]["Tudor"] = 0.40  # edited
    qa["chat_note"] = "Rebalance toward GMT demand"

    inbound = _build_chat_inbound_zip(
        tmp_path / "chat_returned.zip",
        cycle_id=FAKE_CYCLE_ID,
        cycle_focus_bytes=json.dumps(cf).encode("utf-8"),
        monthly_goals_bytes=json.dumps(mg).encode("utf-8"),
        quarterly_allocation_bytes=json.dumps(qa).encode("utf-8"),
    )
    unpack_inbound_bundle(inbound, tmp_path)

    # Edits committed.
    assert json.loads(paths["cycle_focus"].read_text())["chat_commentary"] == (
        "Added SMD Ceramic blue dial based on Q2 momentum"
    )
    assert json.loads(paths["monthly_goals"].read_text())["revenue_target"] == 52000
    assert (
        json.loads(paths["quarterly_allocation"].read_text())["allocations"]["Tudor"]
        == 0.40
    )

    # Unrelated state unchanged.
    assert paths["cache"].read_bytes() == cache_before
    assert paths["ledger"].read_bytes() == ledger_before
    assert paths["brief"].read_bytes() == brief_before


def test_outbound_generates_valid_hashes_for_downstream_consumers(tmp_path):
    """The outbound manifest's sha256 values must themselves verify against
    the archived members — this is what the Chat skill will check on its
    side before accepting an upload."""
    build_fake_grailzee_tree(tmp_path)
    outbound = build_outbound_bundle(tmp_path)
    with zipfile.ZipFile(outbound, "r") as zf:
        manifest = json.loads(zf.read("manifest.json"))
        for entry in manifest["files"]:
            data = zf.read(entry["path"])
            assert _sha256(data) == entry["sha256"], (
                f"role={entry['role']} hash check failed on outbound manifest"
            )
            assert len(data) == entry["size_bytes"]


# ─── Phase 24b: strategy_output.json round trip ───────────────────────


def test_strategy_output_full_round_trip(tmp_path):
    """Simulate a full Chat strategy session handoff:

    1. Agent has a working tree with a cycle cache, prior state files,
       and a sourcing brief (the fixture).
    2. Chat strategy skill produces a strategy_output.json populating
       all four decision sections (cycle_focus + monthly + quarterly +
       one config update).
    3. Operator hands the .json to the plugin.
    4. Plugin validates, atomically writes state/, and archives the
       three operator-facing artifacts to output/briefs/.

    Verifies end to end:
    - All populated decisions land in state/ with the right filenames
    - Non-selected config files are NOT written
    - Archive JSON re-reads byte-equal to the original payload
    - Archive MD matches session_artifacts.cycle_brief_md
    - Archive XLSX carries the Cycle Summary + Config Updates sheets
    - Unrelated state (analysis_cache, trade_ledger, sourcing_brief) is
      untouched
    """
    paths = build_fake_grailzee_tree(tmp_path)
    cache_before = paths["cache"].read_bytes()
    ledger_before = paths["ledger"].read_bytes()
    brief_before = paths["brief"].read_bytes()

    payload = make_strategy_output(
        cycle_id=FAKE_CYCLE_ID,
        session_mode="cycle_planning",
        include_monthly=True,
        include_quarterly=True,
        include_configs=("signal_thresholds",),
        cycle_brief_md="# Cycle Brief\n\nTudor GMT + Omega SMD focus.\n",
    )
    json_path = tmp_path / "strategy_output.json"
    write_strategy_output(json_path, payload)

    result = apply_strategy_output(json_path, paths["root"])

    # State writes: all four populated sections landed.
    assert set(result["roles_written"]) == {
        "cycle_focus",
        "monthly_goals",
        "quarterly_allocation",
        "signal_thresholds",
    }
    assert json.loads(paths["cycle_focus"].read_text()) == payload["decisions"]["cycle_focus"]
    assert json.loads(paths["monthly_goals"].read_text()) == payload["decisions"]["monthly_goals"]
    assert (
        json.loads(paths["quarterly_allocation"].read_text())
        == payload["decisions"]["quarterly_allocation"]
    )
    signal_path = paths["state"] / "signal_thresholds.json"
    assert json.loads(signal_path.read_text()) == (
        payload["decisions"]["config_updates"]["signal_thresholds"]
    )
    # Non-selected configs NOT written.
    for name in (
        "scoring_thresholds",
        "momentum_thresholds",
        "window_config",
        "premium_config",
        "margin_config",
    ):
        assert not (paths["state"] / f"{name}.json").exists()

    # Archive: all three files landed, no errors.
    assert result["archive_errors"] == []
    assert set(result["archive_files_written"]) == {
        f"{FAKE_CYCLE_ID}_strategy_output.json",
        f"{FAKE_CYCLE_ID}_strategy_brief.xlsx",
        f"{FAKE_CYCLE_ID}_strategy_brief.md",
    }
    archived_json = json.loads(
        (paths["briefs"] / f"{FAKE_CYCLE_ID}_strategy_output.json").read_text()
    )
    assert archived_json == payload
    archived_md = (paths["briefs"] / f"{FAKE_CYCLE_ID}_strategy_brief.md").read_text()
    assert archived_md == payload["session_artifacts"]["cycle_brief_md"]
    wb = load_workbook(paths["briefs"] / f"{FAKE_CYCLE_ID}_strategy_brief.xlsx")
    assert "Cycle Summary" in wb.sheetnames
    assert "Config Updates" in wb.sheetnames

    # Unrelated state untouched.
    assert paths["cache"].read_bytes() == cache_before
    assert paths["ledger"].read_bytes() == ledger_before
    assert paths["brief"].read_bytes() == brief_before


def test_strategy_output_partial_update_round_trip(tmp_path):
    """Operator uses a monthly_review session that only updates
    monthly_goals. Other state files (cycle_focus, quarterly) must
    remain untouched at their pre-session contents."""
    paths = build_fake_grailzee_tree(tmp_path)
    cycle_focus_before = paths["cycle_focus"].read_bytes()
    quarterly_before = paths["quarterly_allocation"].read_bytes()

    payload = make_strategy_output(
        cycle_id=FAKE_CYCLE_ID,
        session_mode="monthly_review",
        include_cycle_focus=False,
        include_monthly=True,
    )
    json_path = tmp_path / "strategy_output.json"
    write_strategy_output(json_path, payload)

    result = apply_strategy_output(json_path, paths["root"])

    assert result["roles_written"] == ["monthly_goals"]
    assert result["session_mode"] == "monthly_review"
    assert paths["cycle_focus"].read_bytes() == cycle_focus_before
    assert paths["quarterly_allocation"].read_bytes() == quarterly_before
    assert json.loads(paths["monthly_goals"].read_text()) == (
        payload["decisions"]["monthly_goals"]
    )
