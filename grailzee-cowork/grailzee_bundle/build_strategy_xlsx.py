"""Self-contained XLSX builder for strategy_output payloads.

Produces a human-readable, multi-sheet Vardalux-branded workbook that is
archived alongside state writes when INBOUND applies a strategy_output.json.

Why self-contained
------------------
This module intentionally duplicates the three Vardalux hex color
constants rather than importing them from
``skills/grailzee-eval/scripts/grailzee_common.py``. Rationale:

- The plugin must run from a fresh clone on any machine, any checkout
  path. A hardcoded absolute sys.path into grailzee-eval is non-portable.
- The brand palette doesn't churn. If it grows past three values, the
  right move is to factor out a shared config (e.g.
  ``grailzee-cowork/grailzee_bundle/vardalux_style.py``) consumed by
  both sides — not to cross the plugin/skill boundary with an import.

Canonical source of the color constants remains
``grailzee_common.VARDALUX_COLORS``. If that changes, update here too.

Sheets (each conditional on its decision section being non-null, except
Cycle Summary which is always present):

1. Cycle Summary       — always
2. Targets             — if decisions.cycle_focus is non-null
3. Monthly Goals       — if decisions.monthly_goals is non-null
4. Quarterly           — if decisions.quarterly_allocation is non-null
5. Config Updates      — if decisions.config_updates is non-null
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

# Vardalux brand palette. Hex strings WITHOUT the leading "#", matching
# openpyxl's PatternFill / Font color argument convention. Canonical source:
# skills/grailzee-eval/scripts/grailzee_common.py::VARDALUX_COLORS.
VARDALUX_COLORS: dict[str, str] = {
    "rich_black": "231F20",
    "warm_gold":  "C9A84C",
    "deep_teal":  "315159",
}

WHITE = "FFFFFF"

_SHEET_CYCLE_SUMMARY      = "Cycle Summary"
_SHEET_TARGETS            = "Targets"
_SHEET_MONTHLY_GOALS      = "Monthly Goals"
_SHEET_QUARTERLY          = "Quarterly"
_SHEET_CONFIG_UPDATES     = "Config Updates"


# ─── style helpers ────────────────────────────────────────────────────

def _fill(color_key: str) -> PatternFill:
    return PatternFill(
        start_color=VARDALUX_COLORS[color_key],
        end_color=VARDALUX_COLORS[color_key],
        fill_type="solid",
    )


def _apply_title(cell: Any, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, size=14, color=WHITE)
    cell.fill = _fill("rich_black")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _apply_section_header(cell: Any, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = _fill("deep_teal")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _apply_table_header(cell: Any, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, color=VARDALUX_COLORS["rich_black"])
    cell.fill = _fill("warm_gold")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _apply_label(cell: Any, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, color=VARDALUX_COLORS["deep_teal"])
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws: Worksheet, columns: list[str], widths: list[int]) -> None:
    for col, width in zip(columns, widths):
        ws.column_dimensions[col].width = width


def _format_pct_from_fraction(fraction: float) -> str:
    return f"{fraction * 100:.1f}%"


# ─── sheet builders ───────────────────────────────────────────────────

def _build_cycle_summary_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.create_sheet(_SHEET_CYCLE_SUMMARY)
    cycle_focus = payload["decisions"]["cycle_focus"]

    ws.merge_cells("A1:D1")
    _apply_title(ws["A1"], f"Vardalux Grailzee — {payload['cycle_id']}")

    _apply_label(ws["A3"], "Session Mode")
    ws["B3"] = payload["session_mode"]
    _apply_label(ws["A4"], "Generated At")
    ws["B4"] = payload["generated_at"]
    _apply_label(ws["A5"], "Produced By")
    ws["B5"] = payload["produced_by"]

    if cycle_focus is not None:
        _apply_section_header(ws["A7"], "Cycle Targets")
        ws.merge_cells("A7:D7")

        _apply_label(ws["A8"], "Capital Target")
        ws["B8"] = cycle_focus["capital_target"]
        ws["B8"].number_format = '"$"#,##0'

        _apply_label(ws["A9"], "Volume Target")
        ws["B9"] = cycle_focus["volume_target"]

        _apply_label(ws["A10"], "Target Margin")
        ws["B10"] = _format_pct_from_fraction(cycle_focus["target_margin_fraction"])

        _apply_section_header(ws["A12"], "Brand Emphasis")
        _apply_section_header(ws["C12"], "Brand Pullback")
        for i, brand in enumerate(cycle_focus["brand_emphasis"]):
            ws.cell(row=13 + i, column=1, value=brand)
        for i, brand in enumerate(cycle_focus["brand_pullback"]):
            ws.cell(row=13 + i, column=3, value=brand)

        notes_row = 13 + max(
            len(cycle_focus["brand_emphasis"]), len(cycle_focus["brand_pullback"]), 1
        ) + 1
        _apply_section_header(ws[f"A{notes_row}"], "Notes")
        ws.merge_cells(f"A{notes_row}:D{notes_row}")
        ws[f"A{notes_row + 1}"] = cycle_focus["notes"]
        ws[f"A{notes_row + 1}"].alignment = Alignment(wrap_text=True, vertical="top")

    _autosize(ws, ["A", "B", "C", "D"], [22, 28, 22, 28])


def _build_targets_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.create_sheet(_SHEET_TARGETS)
    cycle_focus = payload["decisions"]["cycle_focus"]

    headers = ["Reference", "Brand", "Model", "Cycle Reason", "MAX BUY Override"]
    for i, header in enumerate(headers, start=1):
        _apply_table_header(ws.cell(row=1, column=i), header)

    for row_idx, target in enumerate(cycle_focus["targets"], start=2):
        ws.cell(row=row_idx, column=1, value=target["reference"])
        ws.cell(row=row_idx, column=2, value=target["brand"])
        ws.cell(row=row_idx, column=3, value=target["model"])
        ws.cell(row=row_idx, column=4, value=target["cycle_reason"])
        override = target["max_buy_override"]
        cell = ws.cell(row=row_idx, column=5, value=override if override is not None else "—")
        if override is not None:
            cell.number_format = '"$"#,##0'

    _autosize(ws, ["A", "B", "C", "D", "E"], [14, 14, 22, 42, 18])


def _build_monthly_goals_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.create_sheet(_SHEET_MONTHLY_GOALS)
    mg = payload["decisions"]["monthly_goals"]

    ws.merge_cells("A1:B1")
    _apply_title(ws["A1"], f"Monthly Goals — {mg['month']}")

    _apply_label(ws["A3"], "Revenue Target")
    ws["B3"] = mg["revenue_target"]
    ws["B3"].number_format = '"$"#,##0'

    _apply_label(ws["A4"], "Volume Target")
    ws["B4"] = mg["volume_target"]

    _apply_section_header(ws["A6"], "Platform Mix")
    ws.merge_cells("A6:B6")

    _apply_table_header(ws["A7"], "Platform")
    _apply_table_header(ws["B7"], "Share %")
    row = 8
    for platform, pct in mg["platform_mix"].items():
        ws.cell(row=row, column=1, value=platform)
        ws.cell(row=row, column=2, value=pct).number_format = '0.0"%"'
        row += 1
    if not mg["platform_mix"]:
        ws.cell(row=row, column=1, value="(not specified)")
        row += 1

    row += 1
    _apply_section_header(ws.cell(row=row, column=1), "Focus Notes")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    ws.cell(row=row, column=1, value=mg["focus_notes"]).alignment = Alignment(
        wrap_text=True, vertical="top"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    row += 2
    _apply_section_header(ws.cell(row=row, column=1), "Review Notes")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    ws.cell(row=row, column=1, value=mg["review_notes"]).alignment = Alignment(
        wrap_text=True, vertical="top"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    _autosize(ws, ["A", "B"], [26, 40])


def _build_quarterly_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.create_sheet(_SHEET_QUARTERLY)
    qa = payload["decisions"]["quarterly_allocation"]

    ws.merge_cells("A1:B1")
    _apply_title(ws["A1"], f"Quarterly Allocation — {qa['quarter']}")

    _apply_section_header(ws["A3"], "Capital Allocation")
    ws.merge_cells("A3:B3")
    _apply_table_header(ws["A4"], "Bucket")
    _apply_table_header(ws["B4"], "Amount")
    row = 5
    for bucket, amount in qa["capital_allocation"].items():
        ws.cell(row=row, column=1, value=bucket)
        ws.cell(row=row, column=2, value=amount).number_format = '"$"#,##0'
        row += 1

    row += 1
    _apply_section_header(ws.cell(row=row, column=1), "Inventory Mix Target")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    _apply_table_header(ws.cell(row=row, column=1), "Tier")
    _apply_table_header(ws.cell(row=row, column=2), "Share %")
    row += 1
    for tier, pct in qa["inventory_mix_target"].items():
        ws.cell(row=row, column=1, value=tier)
        ws.cell(row=row, column=2, value=pct).number_format = '0.0"%"'
        row += 1

    row += 1
    _apply_section_header(ws.cell(row=row, column=1), "Review Notes")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    ws.cell(row=row, column=1, value=qa["review_notes"]).alignment = Alignment(
        wrap_text=True, vertical="top"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    _autosize(ws, ["A", "B"], [26, 40])


def _build_config_updates_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = wb.create_sheet(_SHEET_CONFIG_UPDATES)
    cu = payload["decisions"]["config_updates"]

    ws.merge_cells("A1:C1")
    _apply_title(ws["A1"], "Config Updates")

    _apply_section_header(ws["A3"], "Change Notes")
    ws.merge_cells("A3:C3")
    ws["A4"] = cu["change_notes"]
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A4:C4")

    _apply_table_header(ws["A6"], "Config")
    _apply_table_header(ws["B6"], "Changed?")
    _apply_table_header(ws["C6"], "Envelope Notes")

    sub_keys = (
        "signal_thresholds",
        "scoring_thresholds",
        "momentum_thresholds",
        "window_config",
        "premium_config",
        "margin_config",
    )
    row = 7
    for key in sub_keys:
        sub = cu[key]
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value="Yes" if sub is not None else "No")
        ws.cell(row=row, column=3, value=sub.get("notes", "") if sub else "")
        row += 1

    _autosize(ws, ["A", "B", "C"], [24, 12, 48])


# ─── public entry point ───────────────────────────────────────────────

def build_strategy_xlsx(strategy_output: dict[str, Any], output_path: Path) -> Path:
    """Build the strategy XLSX from a validated strategy_output dict.

    Sheets land in this order: Cycle Summary (always), then Targets,
    Monthly Goals, Quarterly, Config Updates — each included only if its
    backing decisions section is non-null.

    Returns the written path. Creates parent directories as needed.
    """
    wb = Workbook()
    # Workbook() creates a default empty sheet; remove it so sheet order
    # is deterministic and the file has no blank "Sheet1".
    default = wb.active
    if default is not None:
        wb.remove(default)

    _build_cycle_summary_sheet(wb, strategy_output)

    decisions = strategy_output["decisions"]
    if decisions["cycle_focus"] is not None:
        _build_targets_sheet(wb, strategy_output)
    if decisions["monthly_goals"] is not None:
        _build_monthly_goals_sheet(wb, strategy_output)
    if decisions["quarterly_allocation"] is not None:
        _build_quarterly_sheet(wb, strategy_output)
    if decisions["config_updates"] is not None:
        _build_config_updates_sheet(wb, strategy_output)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
